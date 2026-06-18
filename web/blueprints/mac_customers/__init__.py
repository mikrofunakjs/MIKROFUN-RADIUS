"""Customers Blueprint - PPPoE User Management"""
from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify
from web.database import execute_query
from web.decorators import cs_or_admin_required
import datetime

mac_customers_bp = Blueprint('mac_customers', __name__)

@mac_customers_bp.route('/', methods=['GET'])
@cs_or_admin_required
def index():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    q = request.args.get('q', '')
    status_filter = request.args.get('status', '')
    profile_id = request.args.get('profile_id', '')
    
    query = (
        "SELECT c.*, p.name as profile_name, p.price as profile_price, r.name as router_name, "
        "(SELECT COUNT(*) FROM active_sessions s WHERE s.username = c.username) as active_session_count, "
        "(SELECT COALESCE(SUM(acctinputoctets),0) FROM radacct ra WHERE ra.username = c.username) as total_upload, "
        "(SELECT COALESCE(SUM(acctoutputoctets),0) FROM radacct ra WHERE ra.username = c.username) as total_download "
        "FROM customers c "
        "LEFT JOIN profiles p ON c.profile_id = p.id "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE (c.mac_address IS NOT NULL AND c.mac_address != '') "
    )
    params = []
    
    if q:
        query += " AND (c.name LIKE %s OR c.username LIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])
        
    if status_filter:
        query += " AND c.status = %s"
        params.append(status_filter)
        
    if profile_id:
        query += " AND c.profile_id = %s"
        params.append(profile_id)
        
    query += " ORDER BY c.created_at DESC"
    
    customers = execute_query(query, tuple(params), fetch=True)
    
    # Fetch profiles for the dropdown filter (assume they use PPPoE profiles as well)
    profiles = execute_query("SELECT id, name FROM profiles WHERE type='pppoe' ORDER BY name", fetch=True)
    
    # Calculate Statistics
    stats_query = (
        "SELECT "
        "COUNT(*) as total, "
        "SUM(CASE WHEN (SELECT COUNT(*) FROM active_sessions s WHERE s.username = c.username) > 0 THEN 1 ELSE 0 END) as online_count, "
        "SUM(CASE WHEN (SELECT COUNT(*) FROM active_sessions s WHERE s.username = c.username) = 0 AND c.status = 'active' THEN 1 ELSE 0 END) as offline_count, "
        "SUM(CASE WHEN c.status = 'isolir' THEN 1 ELSE 0 END) as isolir_count, "
        "SUM(COALESCE(p.price, 0)) as total_revenue "
        "FROM customers c "
        "LEFT JOIN profiles p ON c.profile_id = p.id "
        "WHERE (c.mac_address IS NOT NULL AND c.mac_address != '')"
    )
    stats = execute_query(stats_query, fetch_one=True) or {'total': 0, 'online_count': 0, 'offline_count': 0, 'isolir_count': 0, 'total_revenue': 0}

    return render_template('mac_customers/list.html', 
                           customers=customers or [], 
                           total_customers=stats['total'],
                           stats=stats,
                           profiles=profiles or [])

@mac_customers_bp.route('/send_bill/<int:id>', methods=['POST'])
@cs_or_admin_required
def send_bill(id):
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    customer = execute_query(
        "SELECT c.*, p.price as profile_price FROM customers c LEFT JOIN profiles p ON c.profile_id = p.id WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    if not customer:
        return jsonify({'success': False, 'message': 'Customer not found.'}), 404

    phone = customer.get('phone')
    if not phone:
        return jsonify({'success': False, 'message': 'User tidak memiliki nomor referensi WA.'}), 400

    # Safely handle billing properties
    try:
        raw_amount = customer.get('profile_price')
        if raw_amount is None:
            raw_amount = 0
            
        amount = '{:,.0f}'.format(raw_amount).replace(',', '.')
        due = customer.get('due_date') or '-'
        status = customer.get('status', 'aktif').upper()
        
        msg = (f"*INFO TAGIHAN MIKROFUN*\n\n"
               f"Yth. Bapak/Ibu *{customer.get('name')}*,\n\n"
               f"Kami informasikan rincian tagihan internet Pelanggan Static IP/MAC Anda:\n"
               f"🔹 *Username:* {customer.get('username')}\n"
               f"🔹 *Tagihan:* Rp {amount}\n"
               f"🔹 *Jatuh Tempo:* {due}\n"
               f"🔹 *Status:* {status}\n\n"
               f"Mohon segera melakukan pembayaran agar layanan tetap aktif. Terima kasih.")

        from web.wa_helper import send_wa
        if send_wa(phone, msg):
            return jsonify({'success': True, 'message': 'Tagihan berhasil dikirim ke antrean.'})
        else:
            return jsonify({'success': False, 'message': 'Gagal mengirim WA.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f"Error Sistem: {str(e)}"}), 500

@mac_customers_bp.route('/bulk_action', methods=['POST'])
def bulk_action():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    action = request.form.get('action')
    customer_ids = request.form.getlist('customer_ids')
    
    if not customer_ids:
        flash('Tidak ada customer yang dipilih!', 'warning')
        return redirect(url_for('mac_customers.index'))
        
    success_count = 0

    if action == 'check_status':
        # Optimized Bulk Check: Group by Router
        format_strings = ','.join(['%s'] * len(customer_ids))
        query = (
             f"SELECT c.username, r.id as router_id, r.vpn_ip, r.api_user, r.api_password, r.api_port "
             f"FROM customers c LEFT JOIN routers r ON c.router_id = r.id "
             f"WHERE c.id IN ({format_strings})"
        )
        customers_to_check = execute_query(query, tuple(customer_ids), fetch=True)
        
        if customers_to_check:
            from web.mikrotik_api import MikrotikApi
            
            # Group by Router
            router_groups = {}
            for c in customers_to_check:
                 if not c['router_id']: continue 
                 if c['router_id'] not in router_groups:
                     router_groups[c['router_id']] = {'creds': c, 'users': []}
                 router_groups[c['router_id']]['users'].append(c['username'])
            
            # Process Each Router
            for rid, group in router_groups.items():
                 creds = group['creds']
                 target_users = set(group['users'])
                 
                 try:
                     api = MikrotikApi(creds['vpn_ip'], creds.get('api_port', 8728))
                     if api.connect() and api.login(creds['api_user'], creds['api_password']):
                         # Fetch ALL active users once
                         active_rows = api.query(['/ppp/active/print'])
                         api.close()
                         
                         active_map = {row.get('name'): row for row in active_rows if row.get('name')}
                         
                         for user in target_users:
                             if user in active_map:
                                 # ONLINE
                                 pkt = active_map[user]
                                 nas_ip = creds['vpn_ip']
                                 session_id = pkt.get('.id', 'unknown')
                                 execute_query(
                                     "INSERT INTO active_sessions (username, nas_ip, acct_session_id) VALUES (%s, %s, %s) "
                                     "ON DUPLICATE KEY UPDATE updated_at=NOW()",
                                     (user, nas_ip, session_id)
                                 )
                             else:
                                 # OFFLINE
                                 execute_query("DELETE FROM active_sessions WHERE username=%s", (user,))
                         
                         success_count += len(target_users)
                 except Exception as e:
                     print(f"Bulk Check Error {creds['vpn_ip']}: {e}")

    else:
        # Standard Item-by-Item Loop
        for cid in customer_ids:
            try:
                if action == 'delete':
                    # Get info for API kick
                    customer = execute_query(
                        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
                        "FROM customers c LEFT JOIN routers r ON c.router_id = r.id WHERE c.id=%s", 
                        (cid,), fetch_one=True
                    )
                    if customer and customer['api_user']:
                        try_api_disconnect(customer, customer['username'])
                    
                    execute_query("DELETE FROM customers WHERE id=%s", (cid,))
                    
                elif action == 'isolir':
                    customer = execute_query(
                        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
                        "FROM customers c LEFT JOIN routers r ON c.router_id = r.id WHERE c.id=%s", 
                        (cid,), fetch_one=True
                    )
                    if customer and customer['api_user']:
                        try_api_disconnect(customer, customer['username'])
                    
                    # Force remove session
                    execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))
                        
                    execute_query("UPDATE customers SET status='isolir' WHERE id=%s", (cid,))
                    
                elif action == 'activate':
                    execute_query("UPDATE customers SET status='active' WHERE id=%s", (cid,))
                    
                success_count += 1
            except Exception as e:
                continue
            
    flash(f'Berhasil memproses {success_count} customer.', 'success')
    return redirect(url_for('mac_customers.index'))

@mac_customers_bp.route('/add', methods=['GET', 'POST'])
@cs_or_admin_required
def add():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    if request.method == 'POST':
        name        = request.form.get('name')
        mac_address = request.form.get('mac_address', '').strip().upper() or None
        static_ip   = request.form.get('static_ip', '').strip() or None
        profile_id  = request.form.get('profile_id') or None
        router_id   = request.form.get('router_id') or None
        phone       = request.form.get('phone', '').strip() or None
        address     = request.form.get('address', '').strip() or None
        due_date    = request.form.get('due_date', '').strip() or None
        odp_id      = request.form.get('odp_id') or None
        port_number = request.form.get('port_number') or None
        coordinates = request.form.get('coordinates', '').strip() or None
        service_type = 'pppoe'

        if not mac_address:
            flash('MAC Address wajib diisi untuk layanan DHCP Statis.', 'error')
            return redirect(url_for('mac_customers.add'))

        if not profile_id:
            flash('Pilih Profile / Paket terlebih dahulu.', 'error')
            return redirect(url_for('mac_customers.add'))

        # For MAC Auth, Username and Password are automatically set to the MAC Address
        username = mac_address
        password = mac_address

        # CHECK LIMIT FOR FREE USERS
        from web.license_service import is_premium
        if not is_premium():
             current_count = execute_query("SELECT COUNT(*) as c FROM customers", fetch_one=True)['c']
             if current_count >= 100:
                  flash('Batas Free: 100 pelanggan. Upgrade ke Premium untuk unlimited.', 'error')
                  return redirect(url_for('mac_customers.add'))

        result = execute_query(
            "INSERT INTO customers (name, username, password, service_type, profile_id, router_id, phone, address, due_date, status, odp_id, port_number, coordinates, mac_address, static_ip) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'active',%s,%s,%s,%s,%s)",
            (name, username, password, service_type, profile_id, router_id, phone, address, due_date, odp_id, port_number, coordinates, mac_address, static_ip)
        )
        if result:
            from web.telegram_helper import send_telegram_message
            send_telegram_message(f"👤 *Pelanggan Baru (MAC/DHCP)*\n\nNama: {name}\nMAC: {mac_address}\nIP Statis: {static_ip or '-'}\nAlamat: {address or '-'}")
            
            from web.blueprints.notifications import add_notification
            add_notification(
                title="Customer Created",
                message=f"Created customer {name} ({username})",
                category="success"
            )
            flash('Customer berhasil dibuat', 'success')
        else:
            flash('Gagal membuat customer', 'error')
            
        return redirect(url_for('mac_customers.index'))
        
    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []
    profiles = execute_query("SELECT * FROM profiles WHERE type=%s ORDER BY name", ('pppoe',), fetch=True) or []
    odps = execute_query("SELECT * FROM odps ORDER BY name", fetch=True) or []
    
    return render_template('mac_customers/add.html', routers=routers, profiles=profiles, odps=odps)

@mac_customers_bp.route('/check_status/<int:id>')
def check_status(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c LEFT JOIN routers r ON c.router_id = r.id WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    
    if not customer:
        flash('Customer tidak ditemukan', 'error')
        return redirect(url_for('mac_customers.index'))
        
    username = customer['username']
    
    # Logic: Connect -> Check. If fail connect/check -> Offline
    try:
        from web.mikrotik_api import MikrotikApi
        
        api = MikrotikApi(customer['vpn_ip'], customer.get('api_port', 8728))
        if api.connect():
             if api.login(customer['api_user'], customer['api_password']):
                 user_data = api.get_active_user(username)
                 api.close()
                 
                 if user_data:
                     # ONLINE
                     nas_ip = customer['vpn_ip']
                     session_id = user_data.get('.id', 'unknown')
                     execute_query(
                         "INSERT INTO active_sessions (username, nas_ip, acct_session_id) VALUES (%s, %s, %s) "
                         "ON DUPLICATE KEY UPDATE updated_at=NOW()",
                         (username, nas_ip, session_id)
                     )
                     flash(f'User {username} is ONLINE', 'success')
                 else:
                     # OFFLINE (User not found in active list)
                     execute_query("DELETE FROM active_sessions WHERE username=%s", (username,))
                     flash(f'User {username} is OFFLINE', 'warning')
             else:
                 api.close()
                 flash('Gagal login ke Mikrotik', 'error')
        else:
             # Connect Fail -> Assume Offline
             execute_query("DELETE FROM active_sessions WHERE username=%s", (username,))
             flash(f'Gagal koneksi ke Mikrotik ({customer["vpn_ip"]}). Status set OFFLINE.', 'error')
             
    except Exception as e:
        flash(f'Error Check Status: {e}', 'error')
        
    return redirect(url_for('mac_customers.index'))

@mac_customers_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@cs_or_admin_required
def edit(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    customer = execute_query("SELECT * FROM customers WHERE id=%s", (id,), fetch_one=True)
    if not customer:
        flash('Customer tidak ditemukan!', 'error')
        return redirect(url_for('mac_customers.index'))

    if request.method == 'POST':
        due_date = request.form.get('due_date') or None
        mac_address = request.form.get('mac_address', '').strip().upper() or None
        static_ip = request.form.get('static_ip', '').strip() or None
        
        if not mac_address:
            flash('MAC Address wajib diisi.', 'error')
            return redirect(url_for('mac_customers.edit', id=id))
            
        # Update username/password to match new MAC address
        username = mac_address
        password = mac_address
        
        execute_query(
            "UPDATE customers SET name=%s, username=%s, password=%s, profile_id=%s, router_id=%s, "
            "phone=%s, address=%s, due_date=%s, odp_id=%s, port_number=%s, coordinates=%s, mac_address=%s, static_ip=%s WHERE id=%s",
            (request.form.get('name'), username, password,
             request.form.get('profile_id'), request.form.get('router_id'),
             request.form.get('phone',''), request.form.get('address',''), due_date,
             request.form.get('odp_id') or None, request.form.get('port_number') or None, 
             request.form.get('coordinates',''), mac_address, static_ip, id)
        )
        flash('Customer berhasil diupdate!', 'success')
        return redirect(url_for('mac_customers.index'))

    profiles = execute_query("SELECT * FROM profiles ORDER BY name", fetch=True) or []
    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []
    odps = execute_query("SELECT * FROM odps ORDER BY name", fetch=True) or []
    return render_template('mac_customers/edit.html', customer=customer, profiles=profiles, routers=routers, odps=odps)

def try_api_disconnect(router_data, username):
    """Helper to disconnect user via Mikrotik API"""
    if not router_data or not router_data.get('vpn_ip') or not router_data.get('api_user'):
        return False, "Router API not configured"
        
    try:
        from web.mikrotik_api import MikrotikApi
        api = MikrotikApi(router_data['vpn_ip'], int(router_data.get('api_port') or 8728))
        
        if not api.login(router_data['api_user'], router_data['api_password']):
            return False, "API Login Failed"
            
        success, msg = api.kick_user(username)
        api.close()
        return success, msg
    except Exception as e:
        return False, str(e)

@mac_customers_bp.route('/delete/<int:id>')
def delete(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    # Get user info and router info BEFORE deleting
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    
    msg = 'Customer berhasil dihapus!'
    
    # 1. API Kick
    if customer and customer['api_user']:
        success, reason = try_api_disconnect(customer, customer['username'])
        if success:
            msg += ' Session diputus (API).'
        else:
            msg += f' (API Kick Warning: {reason})'
    
    # Force remove session
    execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))

    # 2. Delete from DB
    execute_query("DELETE FROM customers WHERE id=%s", (id,))
    
    flash(msg, 'success')
    return redirect(url_for('mac_customers.index'))

@mac_customers_bp.route('/isolir/<int:id>')
def isolir(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    # Get user info
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    
    # 1. Update status in DB
    execute_query("UPDATE customers SET status='isolir' WHERE id=%s", (id,))
    
    msg = 'Customer diisolir!'
    
    # 2. API Kick
    if customer and customer['api_user']:
        success, reason = try_api_disconnect(customer, customer['username'])
        if success:
            msg += ' Session diputus (API).'
        else:
            msg += f' (API Kick Warning: {reason})'
    else:
        msg += ' (Note: API belum disetting di router, user tidak otomatis DC)'
            
    # 3. Force Remove Session from DB (UI Feedback)
    execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))

    flash(msg, 'success')
    return redirect(url_for('mac_customers.index'))

@mac_customers_bp.route('/activate/<int:id>')
def activate(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    execute_query("UPDATE customers SET status='active' WHERE id=%s", (id,))
    flash('Customer diaktifkan!', 'success')
    return redirect(url_for('mac_customers.index'))

@mac_customers_bp.route('/download_template')
@cs_or_admin_required
def download_template():
    """Download blank Excel template for DHCP/MAC bulk import."""
    import io
    from flask import send_file
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        flash('Library openpyxl belum terinstall. Jalankan: pip install openpyxl', 'error')
        return redirect(url_for('mac_customers.index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DHCP Static IP"

    headers = ['name', 'mac_address', 'static_ip', 'profile_name', 'phone', 'address', 'due_date']
    header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Example row
    ws.append(['Sari Wulandari', 'AA:BB:CC:DD:EE:FF', '10.10.10.5', 'Paket 20Mbps', '08129876543', 'Jl. Melati No. 5', '2025-12-31'])

    # Widen columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='template_import_dhcp.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@mac_customers_bp.route('/import', methods=['GET', 'POST'])
@cs_or_admin_required
def import_excel():
    """Import DHCP/Static IP customers from Excel file."""
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    if request.method == 'GET':
        return render_template('mac_customers/import.html')

    try:
        import openpyxl
    except ImportError:
        flash('Library openpyxl belum terinstall. Jalankan: pip install openpyxl', 'error')
        return redirect(url_for('mac_customers.index'))

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('File harus berformat .xlsx (Excel)', 'error')
        return redirect(url_for('mac_customers.import_excel'))

    skip_duplicates = request.form.get('skip_duplicates') == '1'

    results = []
    success_count = 0
    error_count = 0

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        required = {'name', 'mac_address', 'profile_name'}
        if not required.issubset(set(headers)):
            flash(f'Kolom wajib tidak ditemukan. Kolom yang diperlukan: {", ".join(required)}', 'error')
            return redirect(url_for('mac_customers.import_excel'))

        profiles = execute_query("SELECT id, name FROM profiles", fetch=True) or []
        profile_map = {p['name'].lower(): p['id'] for p in profiles}

        import re
        MAC_REGEX = re.compile(r'^([0-9A-Fa-f]{2}[:\-]){5}([0-9A-Fa-f]{2})$')

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            row_data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row) if i < len(headers)}

            name         = row_data.get('name', '')
            mac_address  = row_data.get('mac_address', '').upper().replace('-', ':')
            static_ip    = row_data.get('static_ip', '') or None
            profile_name = row_data.get('profile_name', '')
            phone        = row_data.get('phone', '')
            address      = row_data.get('address', '')
            due_date     = row_data.get('due_date', '') or None

            if not name or not mac_address or not profile_name:
                results.append({'row': row_num, 'name': name or mac_address, 'status': 'error',
                                'message': 'Kolom name, mac_address, profile_name wajib diisi'})
                error_count += 1
                continue

            if not MAC_REGEX.match(mac_address):
                results.append({'row': row_num, 'name': name, 'status': 'error',
                                'message': f'Format MAC Address "{mac_address}" tidak valid (contoh: AA:BB:CC:DD:EE:FF)'})
                error_count += 1
                continue

            profile_id = profile_map.get(profile_name.lower())
            if not profile_id:
                results.append({'row': row_num, 'name': name, 'status': 'error',
                                'message': f'Profile "{profile_name}" tidak ditemukan di database'})
                error_count += 1
                continue

            # MAC address = username & password for RADIUS
            username = mac_address
            password = mac_address

            existing = execute_query("SELECT id FROM customers WHERE username=%s OR mac_address=%s",
                                     (username, mac_address), fetch_one=True)
            if existing:
                if skip_duplicates:
                    results.append({'row': row_num, 'name': name, 'status': 'skip',
                                    'message': f'MAC Address "{mac_address}" sudah ada (dilewati)'})
                    continue
                else:
                    results.append({'row': row_num, 'name': name, 'status': 'error',
                                    'message': f'MAC Address "{mac_address}" sudah ada'})
                    error_count += 1
                    continue

            ok = execute_query(
                "INSERT INTO customers (name, username, password, service_type, profile_id, phone, address, due_date, status, mac_address, static_ip) "
                "VALUES (%s,%s,%s,'pppoe',%s,%s,%s,%s,'active',%s,%s)",
                (name, username, password, profile_id, phone, address, due_date or None, mac_address, static_ip)
            )
            if ok:
                results.append({'row': row_num, 'name': name, 'status': 'success', 'message': 'Berhasil ditambahkan'})
                success_count += 1
            else:
                results.append({'row': row_num, 'name': name, 'status': 'error', 'message': 'Gagal menyimpan ke database'})
                error_count += 1

    except Exception as e:
        flash(f'Gagal memproses file Excel: {e}', 'error')
        return redirect(url_for('mac_customers.import_excel'))

    return render_template('mac_customers/import_result.html',
                           results=results,
                           success_count=success_count,
                           error_count=error_count)
