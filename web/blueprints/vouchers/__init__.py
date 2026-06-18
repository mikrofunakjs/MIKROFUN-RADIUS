from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify, Response
from web.database import execute_query
import random, string, math, io, socket, struct, hashlib, os
from web.decorators import cs_or_admin_required
from web.license_service import is_premium

vouchers_bp = Blueprint('vouchers', __name__)

# ── Code generator ──────────────────────────────────────────────────────────
def generate_code(length=6, prefix='', mode='mixed'):
    if mode == 'numbers':     chars = string.digits
    elif mode == 'upper':     chars = string.ascii_uppercase + string.digits
    elif mode == 'lower':     chars = string.ascii_lowercase + string.digits
    else:                     chars = string.ascii_letters + string.digits
    return f"{prefix}{''.join(random.choice(chars) for _ in range(length))}"


def format_duration(dur):
    if not dur or dur == 0:
        return "Unlimited"
    dur = float(dur)
    if dur >= 720 and dur % 720 == 0:
        return f"{int(dur // 720)} Bulan"
    if dur >= 24 and dur % 24 == 0:
        return f"{int(dur // 24)} Hari"
    return f"{int(dur) if dur == int(dur) else dur} Jam"


# ─────────────────────────────────────────────────────────────────────────────
# RADIUS CoA HELPER (Disconnect-Request – RFC 3576)
# ─────────────────────────────────────────────────────────────────────────────
def _send_coa_disconnect(nas_ip: str, nas_secret: str, session_id: str,
                         username: str = '', nas_port: int = 1700) -> bool:
    """
    Send a RADIUS Disconnect-Request (CoA) to a NAS to terminate a session.
    Mikrotik uses port 1700 by default (Radius > Incoming).
    Returns True on success (Disconnect-ACK), False otherwise.
    """
    try:
        # Build RADIUS packet
        code       = 40   # Disconnect-Request
        identifier = int.from_bytes(os.urandom(1), 'big')
        attrs      = b''

        def add_attr(attr_type, value_bytes):
            return bytes([attr_type, len(value_bytes) + 2]) + value_bytes

        if username:
            attrs += add_attr(1, username.encode())    # User-Name
        if session_id:
            attrs += add_attr(44, session_id.encode()) # Acct-Session-Id

        length = 20 + len(attrs)
        # Authenticator: MD5(Code + ID + Length + 16x\x00 + Attrs + secret)
        auth_placeholder = b'\x00' * 16
        pkt = bytes([code, identifier]) + struct.pack('!H', length) + auth_placeholder + attrs
        authenticator = hashlib.md5(pkt + nas_secret.encode()).digest()
        pkt = bytes([code, identifier]) + struct.pack('!H', length) + authenticator + attrs

        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.settimeout(3)
        sock.sendto(pkt, (nas_ip, nas_port))

        data, _ = sock.recvfrom(4096)
        sock.close()
        # Code 41 = Disconnect-ACK (success), 42 = Disconnect-NAK (fail)
        if data[0] == 41:
            return True
        elif data[0] == 42:
            # Maybe log Error-Cause (Attr 101)?
            return False
        return False
    except Exception as e:
        print(f"[CoA Error] {nas_ip}:{nas_port} -> {e}")
        return False





# ── INDEX: list with filter/search/pagination ────────────────────────────────
@vouchers_bp.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    # — filters —
    q            = request.args.get('q', '').strip()
    status_f     = request.args.get('status', '')
    profile_f    = request.args.get('profile_id', '')
    batch_f      = request.args.get('batch_id', '')
    page         = max(1, int(request.args.get('page', 1)))
    per_page     = 50
    offset       = (page - 1) * per_page

    where = "WHERE 1=1"
    params = []
    if q:
        where += " AND v.code LIKE %s"
        params.append(f"%{q}%")
    if status_f:
        where += " AND v.status=%s"
        params.append(status_f)
    if profile_f:
        where += " AND v.profile_id=%s"
        params.append(profile_f)
    if batch_f:
        where += " AND v.batch_id=%s"
        params.append(batch_f)

    base_q = (
        "FROM vouchers v "
        "LEFT JOIN profiles p ON v.profile_id=p.id "
        "LEFT JOIN voucher_batches b ON v.batch_id=b.id "
        + where
    )
    count  = (execute_query(f"SELECT COUNT(*) as c {base_q}", tuple(params), fetch_one=True) or {}).get('c', 0)
    vouchers = execute_query(
        f"SELECT v.*, p.name as profile_name, p.validity, p.validity_unit, b.name as batch_name {base_q} "
        f"ORDER BY v.id DESC LIMIT %s OFFSET %s",
        tuple(params) + (per_page, offset), fetch=True
    ) or []

    total_pages = max(1, math.ceil(count / per_page))

    # — sidebar data —
    profiles = execute_query("SELECT id, name FROM profiles WHERE type='voucher' ORDER BY name", fetch=True) or []
    batches  = execute_query("SELECT id, name FROM voucher_batches ORDER BY created_at DESC LIMIT 50", fetch=True) or []

    # — quick stats —
    stats    = execute_query(
        "SELECT status, COUNT(*) as c FROM vouchers GROUP BY status", fetch=True
    ) or []
    stats_map = {s['status']: s['c'] for s in stats}

    total_vouchers = execute_query("SELECT COUNT(*) as c FROM vouchers", fetch_one=True)['c']

    # Enrich with human readable validity
    enriched = []
    for v in vouchers:
        dur = v.get('duration_hours')
        if dur is None:
            # Fallback to profile validity
            val = v.get('validity', 24)
            unit = v.get('validity_unit', 'hours')
            dur = val
            if unit == 'days': dur = val * 24
            elif unit == 'months': dur = val * 720
            
        v_dur_str = format_duration(dur)
        enriched.append({**v, 'validity_str': v_dur_str})

    return render_template('vouchers/list.html',
                           vouchers=enriched, count=count, total_pages=total_pages, page=page,
                           q=q, status_f=status_f, profile_f=profile_f, batch_f=batch_f,
                           profiles=profiles, batches=batches, stats_map=stats_map,
                           total_vouchers=total_vouchers)


# ── STATS DASHBOARD ──────────────────────────────────────────────────────────
@vouchers_bp.route('/stats')
def stats():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    summary = execute_query(
        "SELECT status, COUNT(*) as c, SUM(price) as total FROM vouchers GROUP BY status", fetch=True
    ) or []
    by_profile = execute_query(
        "SELECT p.name, COUNT(v.id) as c, "
        "SUM(v.status='used') as used_c, SUM(v.price) as revenue "
        "FROM vouchers v LEFT JOIN profiles p ON v.profile_id=p.id "
        "GROUP BY v.profile_id ORDER BY used_c DESC", fetch=True
    ) or []
    by_date = execute_query(
        "SELECT DATE(created_at) as day, COUNT(*) as c, SUM(status='used') as used_c "
        "FROM vouchers WHERE created_at >= DATE_SUB(NOW(), INTERVAL 30 DAY) "
        "GROUP BY DATE(created_at) ORDER BY day DESC", fetch=True
    ) or []
    batches = execute_query(
        "SELECT b.*, COUNT(v.id) as total_v, SUM(v.status='unused') as unused_v "
        "FROM voucher_batches b LEFT JOIN vouchers v ON v.batch_id=b.id "
        "GROUP BY b.id ORDER BY b.created_at DESC LIMIT 20", fetch=True
    ) or []
    return render_template('vouchers/stats.html', summary=summary, by_profile=by_profile,
                           by_date=by_date, batches=batches)


# ── ADD / GENERATE ───────────────────────────────────────────────────────────
@vouchers_bp.route('/add', methods=['GET', 'POST'])
@cs_or_admin_required
def add():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    profiles = execute_query("SELECT * FROM profiles WHERE type='voucher' ORDER BY name", fetch=True) or []

    if request.method == 'POST':
        qty        = int(request.form.get('qty', 1))
        profile_id = request.form.get('profile_id')
        prefix     = request.form.get('prefix', '')
        length     = int(request.form.get('length', 6))
        mode       = request.form.get('mode', 'mixed')
        batch_name = request.form.get('batch_name', '').strip()
        expires_at = request.form.get('expires_at', '') or None

        profile = execute_query("SELECT * FROM profiles WHERE id=%s", (profile_id,), fetch_one=True)
        if not profile:
            flash('Profile tidak valid', 'error')
            return redirect(url_for('vouchers.add'))

        if not is_premium():
            current_count = execute_query("SELECT COUNT(*) as c FROM vouchers", fetch_one=True)['c']
            if current_count + qty > 800:
                flash('Batas Free: 800 voucher. Upgrade ke Premium untuk unlimited.', 'error')
                return redirect(url_for('vouchers.add'))

        # Create / reuse batch
        batch_id = None
        if batch_name:
            execute_query("INSERT INTO voucher_batches (name, profile_id, created_at) VALUES (%s,%s,NOW())",
                          (batch_name, profile_id))
            row = execute_query("SELECT id FROM voucher_batches WHERE name=%s ORDER BY id DESC LIMIT 1",
                                (batch_name,), fetch_one=True)
            if row:
                batch_id = row['id']

        # Calculate duration in hours
        validity = int(profile.get('validity', 24))
        unit = profile.get('validity_unit', 'hours')
        
        duration_hours = validity
        if unit == 'days':
            duration_hours = validity * 24
        elif unit == 'months':
            duration_hours = validity * 720 # 30 days

        success_count = 0
        generated_ids = []
        for _ in range(qty):
            code = generate_code(length, prefix, mode)
            try:
                execute_query(
                    "INSERT INTO vouchers (code, profile_id, duration_hours, price, quota_limit, status, created_by, batch_id, expires_at) "
                    "VALUES (%s,%s,%s,%s,%s,'unused','admin',%s,%s)",
                    (code, profile_id, duration_hours, profile.get('price', 0),
                     profile.get('quota_limit', 0), batch_id, expires_at)
                )
                generated_ids.append(code)
                success_count += 1
            except Exception:
                continue

        flash(f'Berhasil generate {success_count} voucher.', 'success')
        from web.blueprints.notifications import add_notification
        add_notification("Voucher Generated",
                         f"Admin generated {success_count} vouchers (batch: {batch_name or '-'}).",
                         "success")
        # ── FINANCE LEDGER: catat pemasukan voucher admin ──────────────────
        try:
            from web.finance_helper import record_admin_voucher
            record_admin_voucher(success_count, profile, batch_name,
                                 recorded_by=session.get('username', 'admin'))
        except Exception as _fe:
            print(f"[finance] warn: {_fe}")
        # ──────────────────────────────────────────────────────────────────
        codes_str = ",".join(generated_ids)
        return redirect(url_for('vouchers.print_batch', codes=codes_str))

    return render_template('vouchers/add.html', profiles=profiles)


# ── DELETE SINGLE ────────────────────────────────────────────────────────────
@vouchers_bp.route('/delete/<int:id>')
def delete(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    execute_query("DELETE FROM vouchers WHERE id=%s", (id,))
    flash('Voucher dihapus', 'success')
    return redirect(url_for('vouchers.index'))


# ── BULK DELETE ───────────────────────────────────────────────────────────────
@vouchers_bp.route('/bulk_delete', methods=['POST'])
def bulk_delete():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    mode = request.form.get('mode', '')  # 'used', 'expired', 'selected', 'batch'
    batch_id = request.form.get('batch_id', '')
    ids = request.form.getlist('ids')

    if mode == 'used':
        execute_query("DELETE FROM vouchers WHERE status='used'")
        flash('Semua voucher yang sudah dipakai berhasil dihapus.', 'success')
    elif mode == 'expired':
        execute_query("DELETE FROM vouchers WHERE status='expired' OR (expires_at IS NOT NULL AND expires_at < NOW() AND status='unused')")
        flash('Semua voucher expired berhasil dihapus.', 'success')
    elif mode == 'batch' and batch_id:
        execute_query("DELETE FROM vouchers WHERE batch_id=%s", (batch_id,))
        execute_query("DELETE FROM voucher_batches WHERE id=%s", (batch_id,))
        flash('Batch voucher berhasil dihapus.', 'success')
    elif mode == 'selected' and ids:
        placeholders = ','.join(['%s'] * len(ids))
        execute_query(f"DELETE FROM vouchers WHERE id IN ({placeholders})", tuple(ids))
        flash(f'{len(ids)} voucher berhasil dihapus.', 'success')
    else:
        flash('Tidak ada voucher yang dihapus.', 'warning')

    return redirect(url_for('vouchers.index'))


# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────
@vouchers_bp.route('/export')
def export():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    status_f  = request.args.get('status', '')
    batch_f   = request.args.get('batch_id', '')
    profile_f = request.args.get('profile_id', '')

    where = "WHERE 1=1"
    params = []
    if status_f:
        where += " AND v.status=%s"; params.append(status_f)
    if batch_f:
        where += " AND v.batch_id=%s"; params.append(batch_f)
    if profile_f:
        where += " AND v.profile_id=%s"; params.append(profile_f)

    vouchers = execute_query(
        "SELECT v.code, v.status, v.price, p.name as profile, v.duration_hours, "
        "v.expires_at, v.created_at, b.name as batch "
        "FROM vouchers v "
        "LEFT JOIN profiles p ON v.profile_id=p.id "
        "LEFT JOIN voucher_batches b ON v.batch_id=b.id "
        + where + " ORDER BY v.id DESC", tuple(params), fetch=True
    ) or []

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vouchers"
        headers = ['Kode', 'Status', 'Profile', 'Durasi (Jam)', 'Harga', 'Batch', 'Expires At', 'Created At']
        ws.append(headers)
        for v in vouchers:
            ws.append([
                v.get('code'), v.get('status'), v.get('profile'),
                v.get('duration_hours'), float(v.get('price') or 0),
                v.get('batch'), str(v.get('expires_at') or ''), str(v.get('created_at') or '')
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=vouchers.xlsx'}
        )
    except ImportError:
        flash('openpyxl tidak tersedia di server ini.', 'error')
        return redirect(url_for('vouchers.index'))


# ── PRINT BATCH : selective via checkboxes ───────────────────────────────────
@vouchers_bp.route('/print_batch', methods=['GET'])
def print_batch():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    codes   = request.args.get('codes')
    limit   = request.args.get('limit', 5000)
    batch   = request.args.get('batch_id', '')
    profile = request.args.get('profile_id', '')

    if codes:
        code_list    = codes.split(',')
        placeholders = ','.join(['%s'] * len(code_list))
        vouchers = execute_query(
            f"SELECT v.*, p.name as profile_name, p.validity, p.validity_unit "
            f"FROM vouchers v LEFT JOIN profiles p ON v.profile_id=p.id "
            f"WHERE v.code IN ({placeholders}) ORDER BY v.code",
            tuple(code_list), fetch=True
        ) or []
    elif batch:
        vouchers = execute_query(
            "SELECT v.*, p.name as profile_name, p.validity, p.validity_unit "
            "FROM vouchers v LEFT JOIN profiles p ON v.profile_id=p.id "
            "WHERE v.batch_id=%s ORDER BY v.code",
            (batch,), fetch=True
        ) or []
    else:
        where_clause = "WHERE v.status='unused'"
        params = []
        if profile:
            where_clause += " AND v.profile_id=%s"
            params.append(profile)

        vouchers = execute_query(
            f"SELECT v.*, p.name as profile_name, p.validity, p.validity_unit "
            f"FROM vouchers v LEFT JOIN profiles p ON v.profile_id=p.id "
            f"{where_clause} ORDER BY v.created_at DESC LIMIT %s",
            tuple(params) + (int(limit),), fetch=True
        ) or []
        
    # Enrich
    for v in vouchers:
        dur = v.get('duration_hours')
        if dur is None:
            val = v.get('validity', 24)
            unit = v.get('validity_unit', 'hours')
            dur = val
            if unit == 'days': dur = val * 24
            elif unit == 'months': dur = val * 720
        v['validity_str'] = format_duration(dur)

    return render_template('vouchers/print.html', vouchers=vouchers)


# ── PRINT BY IDs (for checkbox-selected print) ───────────────────────────────
@vouchers_bp.route('/print_batch_ids')
def print_batch_ids():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    ids_str = request.args.get('ids', '')
    if not ids_str:
        return redirect(url_for('vouchers.index'))
    ids = [i for i in ids_str.split(',') if i.isdigit()]
    if not ids:
        return redirect(url_for('vouchers.index'))
    placeholders = ','.join(['%s'] * len(ids))
    vouchers = execute_query(
        f"SELECT v.*, p.name as profile_name, p.validity, p.validity_unit "
        f"FROM vouchers v LEFT JOIN profiles p ON v.profile_id=p.id "
        f"WHERE v.id IN ({placeholders}) ORDER BY v.id",
        tuple(ids), fetch=True
    ) or []
    
    # Enrich
    for v in vouchers:
        dur = v.get('duration_hours')
        if dur is None:
            val = v.get('validity', 24)
            unit = v.get('validity_unit', 'hours')
            dur = val
            if unit == 'days': dur = val * 24
            elif unit == 'months': dur = val * 720
        v['validity_str'] = format_duration(dur)
        
    return render_template('vouchers/print.html', vouchers=vouchers)


# ── HISTORY ───────────────────────────────────────────────────────────────────
@vouchers_bp.route('/history')
def history():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    transactions = execute_query(
        "SELECT p.*, c.name as customer_name, pr.name as profile_name "
        "FROM payments p LEFT JOIN customers c ON p.customer_id=c.id "
        "LEFT JOIN profiles pr ON p.profile_id=pr.id "
        "WHERE p.payment_type='voucher' ORDER BY p.created_at DESC",
        fetch=True
    ) or []
    return render_template('vouchers/history.html', transactions=transactions)


@vouchers_bp.route('/history/approve/<int:payment_id>', methods=['POST'])
def history_approve(payment_id):
    if not session.get('logged_in'): return redirect(url_for('auth.login'))
    payment = execute_query(
        "SELECT * FROM payments WHERE id=%s AND payment_type='voucher' AND status='pending'",
        (payment_id,), fetch_one=True
    )
    if not payment:
        flash("Transaksi tidak valid atau sudah diproses.", "error")
        return redirect(url_for('vouchers.history'))
    from web.blueprints.api import generate_voucher_for_payment
    try:
        generate_voucher_for_payment(payment)
        execute_query("UPDATE payments SET status='approved' WHERE id=%s", (payment_id,))
        flash("Manual Approve Berhasil! Voucher telah di-generate.", "success")
    except Exception as e:
        flash(f"Error Generate Voucher: {e}", "error")
    return redirect(url_for('vouchers.history'))


# ── ACTIVE VOUCHER MONITOR ────────────────────────────────────────────────────
@vouchers_bp.route('/active')
def active():
    """Show all currently active (in-use) vouchers with remaining time."""
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    vouchers = execute_query(
        "SELECT v.*, p.name as profile_name, p.validity, p.validity_unit, "
        "COALESCE(r_nas.name, r_prof.name) as router_name, "
        "COALESCE(r_nas.vpn_ip, r_prof.vpn_ip) as router_ip "
        "FROM vouchers v "
        "LEFT JOIN profiles p ON v.profile_id = p.id "
        "LEFT JOIN routers r_nas ON v.nas_id = r_nas.id "
        "LEFT JOIN routers r_prof ON p.router_id = r_prof.id "
        "WHERE v.status='active' "
        "ORDER BY v.activated_at DESC",
        fetch=True
    ) or []

    # Build per-voucher remaining seconds for the template
    import datetime
    now = datetime.datetime.now()
    enriched = []
    for v in vouchers:
        remaining_sec = None
        expires_ts = None
        if v.get('activated_at') and v.get('duration_hours'):
            exp = v['activated_at'] + datetime.timedelta(hours=float(v['duration_hours']))
            expires_ts = exp.strftime('%Y-%m-%dT%H:%M:%S')
            delta = exp - now
            remaining_sec = max(0, int(delta.total_seconds()))
        
        # Format Human Readable Duration
        dur = v.get('duration_hours')
        if dur is None:
            val = v.get('validity', 24)
            unit = v.get('validity_unit', 'hours')
            dur = val
            if unit == 'days': dur = val * 24
            elif unit == 'months': dur = val * 720
            
        v_dur_str = format_duration(dur)

        enriched.append({**v, 'remaining_sec': remaining_sec, 'expires_ts': expires_ts, 'validity_str': v_dur_str})

    # Extract unique routers for the dropdown filter
    active_routers = {}
    for v in enriched:
        if v.get('router_name'):
            active_routers[v['router_name']] = v['router_ip']

    return render_template('vouchers/active.html', vouchers=enriched, active_routers=active_routers)


# ── API: Active vouchers JSON refresh ─────────────────────────────────────────
@vouchers_bp.route('/active/api')
def active_api():
    """JSON endpoint for auto-refresh of active vouchers."""
    if not session.get('logged_in'):
        return jsonify([])
    import datetime
    rows = execute_query(
        "SELECT v.id, v.code, v.duration_hours, v.activated_at, v.session_id, "
        "v.nas_id, p.name as profile_name, "
        "COALESCE(r_nas.vpn_ip, r_prof.vpn_ip) as router_ip, "
        "COALESCE(r_nas.name, r_prof.name) as router_name "
        "FROM vouchers v "
        "LEFT JOIN profiles p ON v.profile_id = p.id "
        "LEFT JOIN routers r_nas ON v.nas_id = r_nas.id "
        "LEFT JOIN routers r_prof ON p.router_id = r_prof.id "
        "WHERE v.status='active' ORDER BY v.activated_at DESC",
        fetch=True
    ) or []
    now = datetime.datetime.now()
    result = []
    for v in rows:
        remaining = None
        if v.get('activated_at') and v.get('duration_hours'):
            exp = v['activated_at'] + datetime.timedelta(hours=float(v['duration_hours']))
            remaining = max(0, int((exp - now).total_seconds()))
        result.append({'id': v['id'], 'code': v['code'],
                       'profile': v['profile_name'], 'remaining': remaining})
    return jsonify(result)


# ── API: Sync active users from MikroTik ────────────────────────────────────────
@vouchers_bp.route('/active/mikrotik')
def active_mikrotik():
    """Pull real-time active hotspot users from ALL online MikroTik routers"""
    if not session.get('logged_in'):
        return jsonify([])

    from web.mikrotik_api import MikrotikApi
    routers = execute_query("SELECT * FROM routers WHERE status='online'", fetch=True) or []
    
    all_users = []
    for router in routers:
        try:
            api = MikrotikApi(router.get('vpn_ip', router.get('ip_address', '')))
            api.username = router.get('api_user', 'admin')
            api.password = router.get('api_password', '')
            active = api.get_hotspot_active()
            for u in active:
                u['router_name'] = router.get('name', 'Unknown')
                u['router_ip'] = router.get('vpn_ip', router.get('ip_address', ''))
            all_users.extend(active)
        except Exception as e:
            print(f"[MT Active] {router.get('name')} error: {e}")

    return jsonify(all_users)


# ── SOFT DISCONNECT (expire) ──────────────────────────────────────────────────
@vouchers_bp.route('/disconnect/<int:id>', methods=['POST'])
def disconnect(id):
    """Soft disconnect: mark voucher as expired (RADIUS will reject next auth)."""
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    v = execute_query("SELECT * FROM vouchers WHERE id=%s", (id,), fetch_one=True)
    if not v:
        flash('Voucher tidak ditemukan.', 'error')
        return redirect(url_for('vouchers.active'))
    execute_query("UPDATE vouchers SET status='expired' WHERE id=%s", (id,))
    flash(f'Voucher {v["code"]} dinonaktifkan (soft). Session putus saat RADIUS check berikutnya.', 'warning')
    return redirect(url_for('vouchers.active'))


# ── HARD DISCONNECT (CoA Disconnect-Request) ──────────────────────────────────
@vouchers_bp.route('/coa_disconnect/<int:id>', methods=['POST'])
def coa_disconnect(id):
    """Hard disconnect: send RADIUS CoA Disconnect-Request to router NAS."""
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    v = execute_query(
        "SELECT v.*, "
        "COALESCE(r_nas.vpn_ip, r_prof.vpn_ip) as router_ip, "
        "COALESCE(r_nas.api_user, r_prof.api_user) as api_user, "
        "COALESCE(r_nas.api_password, r_prof.api_password) as api_password, "
        "COALESCE(r_nas.api_port, r_prof.api_port) as api_port "
        "FROM vouchers v "
        "LEFT JOIN profiles p ON v.profile_id = p.id "
        "LEFT JOIN routers r_nas ON v.nas_id = r_nas.id "
        "LEFT JOIN routers r_prof ON p.router_id = r_prof.id "
        "WHERE v.id=%s", (id,), fetch_one=True
    )
    if not v:
        flash('Voucher tidak ditemukan.', 'error')
        return redirect(url_for('vouchers.active'))

    nas_ip = v.get('router_ip')
    
    # Auto-fetch RADIUS secret from settings (default: testing123)
    secret_row = execute_query("SELECT setting_value FROM settings WHERE setting_key='radius_secret'", fetch_one=True)
    nas_secret = secret_row['setting_value'] if secret_row and secret_row.get('setting_value') else 'testing123'

    session_id = v.get('session_id') or ''
    username   = v.get('code') or ''

    # --- FALLBACK: Search in active_sessions table for MAC Address and NAS IP ---
    mac_address = None
    active_sess = execute_query(
        "SELECT * FROM active_sessions WHERE username=%s ORDER BY id DESC LIMIT 1",
        (username,), fetch_one=True
    )
    if active_sess:
        # FIX: Correct column name is 'mac_address'
        mac_address = active_sess.get('mac_address')
        if not v.get('nas_id') or not nas_ip:
            nas_ip = active_sess['nas_ip']
            session_id = active_sess['acct_session_id']
            # Re-fetch router info based on this nas_ip
            r_fallback = None
            try:
                r_fallback = execute_query(
                    "SELECT * FROM routers WHERE ip_address=%s OR vpn_ip=%s OR vpn_ip LIKE %s LIMIT 1",
                    (nas_ip, nas_ip, f"%{nas_ip}%"), fetch_one=True
                )
            except:
                # Fallback if ip_address column doesn't exist yet
                r_fallback = execute_query(
                    "SELECT * FROM routers WHERE vpn_ip=%s OR vpn_ip LIKE %s LIMIT 1",
                    (nas_ip, f"%{nas_ip}%"), fetch_one=True
                )
            if r_fallback:
                # Prioritize VPN IP for reachability if it exists
                if r_fallback.get('vpn_ip'):
                    nas_ip = r_fallback['vpn_ip']
                v['api_user'] = r_fallback.get('api_user')
                v['api_password'] = r_fallback.get('api_password')
                v['api_port'] = r_fallback.get('api_port')

    # Always soft-expire first (fallback safety)
    execute_query("UPDATE vouchers SET status='expired' WHERE id=%s", (id,))

    if nas_ip:
        # 1. Try RADIUS CoA on port 1700 (Mikrotik standard)
        coa_ok = _send_coa_disconnect(nas_ip, nas_secret, session_id, username, nas_port=1700)
        
        # 2. Try RADIUS CoA on port 3799 (RFC standard fallback) if 1700 failed
        if not coa_ok:
            coa_ok = _send_coa_disconnect(nas_ip, nas_secret, session_id, username, nas_port=3799)
        
        # 3. Try Mikrotik API guarantee drop as primary backup
        api_ok, api_msg = False, "No API Credentials"
        if v.get('api_user') and v.get('api_password'):
            from web.mikrotik_api import MikrotikApi
            
            # Note: We must use vpn_ip/ip_address. If nas_ip is inaccessible directly 
            # (e.g. private NAT), the API connection will timeout.
            try:
                api = MikrotikApi(nas_ip, int(v.get('api_port', 8728)), timeout=5)
                if api.login(v['api_user'], v['api_password']):
                    api_ok, api_msg = api.kick_hotspot_user(username, mac_address)
                    api.close()
            except Exception as e:
                print(f"Disconnect Mikrotik API Error: {e}")
        
        if coa_ok or api_ok:
            flash(f'✅ Disconnect berhasil! Sesi voucher {username} diputus di router.', 'success')
        else:
            flash(f'⚠️ Router tidak merespon perintah putus (Port 1700/3799/API). Pastikan "Radius > Incoming > Accept" aktif di Mikrotik.', 'warning')
    else:
        flash(f'Voucher {username} dinonaktifkan (soft). Tidak ada info router.', 'warning')

    return redirect(url_for('vouchers.active'))

@vouchers_bp.route('/import', methods=['GET', 'POST'])
def import_excel():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Tidak ada file yang dipilih.', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Nama file kosong.', 'error')
            return redirect(request.url)
            
        if file and file.filename.endswith('.xlsx'):
            import os
            from openpyxl import load_workbook
            from web.app import app
            
            # CHECK LIMIT
            if not is_premium():
                c = execute_query("SELECT COUNT(*) as c FROM vouchers", fetch_one=True)['c']
                if c >= 800:
                    flash('Batas Free: 800 voucher. Hapus beberapa atau upgrade ke Premium.', 'error')
                    return redirect(request.url)
            
            # Use app config for upload folder
            upload_dir = app.config.get('UPLOAD_FOLDER', os.path.join(os.getcwd(), 'web', 'static', 'uploads'))
            if not os.path.exists(upload_dir): os.makedirs(upload_dir)
            
            filepath = os.path.join(upload_dir, file.filename)
            file.save(filepath)
            
            try:
                wb = load_workbook(filepath)
                sheet = wb.active
                
                # Header Mapping
                headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
                
                count = 0
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row) if i < len(headers)}
                        
                        code = row_data.get('code')
                        if not code: continue
                        
                        profile_name = row_data.get('profile_name')
                        batch_name = row_data.get('batch_name')
                        
                        # Type validation (numeric hardening)
                        def to_num(val, default=0):
                            try:
                                return float(val) if val else default
                            except:
                                return default

                        price = to_num(row_data.get('price'))
                        buy_price = to_num(row_data.get('buy_price'))
                        duration = to_num(row_data.get('duration'))
                        if not duration: duration = to_num(row_data.get('duration_hours'))
                        
                        expires = row_data.get('expires', None)
                        if not expires: expires = row_data.get('expires_at', None)
                        if str(expires).strip() == '': expires = None
                        
                        profile_id = None
                        if profile_name:
                            p = execute_query("SELECT id FROM profiles WHERE name=%s", (profile_name,), fetch_one=True)
                            if p: profile_id = p['id']
                        
                        batch_id = None
                        if batch_name:
                            b = execute_query("SELECT id FROM voucher_batches WHERE name=%s", (batch_name,), fetch_one=True)
                            if not b:
                                execute_query("INSERT INTO voucher_batches (name, profile_id, created_at) VALUES (%s,%s,NOW())", (batch_name, profile_id))
                                b = execute_query("SELECT id FROM voucher_batches WHERE name=%s", (batch_name,), fetch_one=True)
                            if b: batch_id = b['id']

                        execute_query(
                            "INSERT INTO vouchers (code, profile_id, batch_id, price, buy_price, duration_hours, expires_at, status, created_by) "
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, 'unused', 'admin') "
                            "ON DUPLICATE KEY UPDATE profile_id=%s, batch_id=%s, price=%s, buy_price=%s, duration_hours=%s, expires_at=%s",
                            (code, profile_id, batch_id, price, buy_price, duration, expires,
                             profile_id, batch_id, price, buy_price, duration, expires)
                        )
                        count += 1
                    except Exception as row_err:
                        print(f"Row {row_idx} Error: {row_err}")
                        continue
                
                flash(f'Berhasil mengimport {count} Voucher.', 'success')
                return redirect(url_for('vouchers.index'))
            except Exception as e:
                flash(f'Gagal membaca Excel: {e}', 'error')
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
        else:
            flash('File tidak valid. Harap upload .xlsx', 'error')
            
    return render_template('vouchers/import.html')
