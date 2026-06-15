from flask import Blueprint, render_template, request, session, redirect, url_for, flash
from web.database import execute_query
from web.decorators import admin_required

odp_bp = Blueprint('odp', __name__)

@odp_bp.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    # Get ODPs with customer count
    odps = execute_query(
        "SELECT o.*, COUNT(c.id) as used_ports "
        "FROM odps o "
        "LEFT JOIN customers c ON o.id = c.odp_id "
        "GROUP BY o.id "
        "ORDER BY o.name ASC",
        fetch=True
    ) or []
    
    return render_template('odp/index.html', odps=odps)

@odp_bp.route('/map')
def map_view():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    # Get all ODPs with usage stats and backbone parent
    odps = execute_query(
        "SELECT o.*, COUNT(c.id) as used_ports "
        "FROM odps o "
        "LEFT JOIN customers c ON o.id = c.odp_id "
        "GROUP BY o.id", 
        fetch=True
    ) or []
    
    # Get all Customers with coordinates - use simple query first, then enrich
    customers = []
    try:
        # Fetch all customers who have coordinates populated
        raw_customers = execute_query(
            "SELECT c.id, c.name, c.username, c.service_type, "
            "c.coordinates, c.odp_id, c.address "
            "FROM customers c "
            "WHERE c.coordinates IS NOT NULL",
            fetch=True
        ) or []
        
        # Filter valid coordinates in python
        for c in raw_customers:
            coord = str(c.get('coordinates', '')).strip()
            if coord and coord.lower() != 'none' and ',' in coord:
                c['coordinates'] = coord
                customers.append(c)
    except Exception as e:
        print(f"[MAP] Customer query error: {e}")
        customers = []
    
    # Enrich with online status (separate query to avoid breaking main data)
    for c in customers:
        c['is_online'] = 0
        c['open_tickets'] = 0
        try:
            sess = execute_query(
                "SELECT COUNT(*) as cnt FROM active_sessions WHERE username=%s",
                (c['username'],), fetch_one=True
            )
            if sess:
                c['is_online'] = sess['cnt']
        except:
            pass
        try:
            tkt = execute_query(
                "SELECT COUNT(*) as cnt FROM tickets WHERE customer_id=%s AND status='open'",
                (c['id'],), fetch_one=True
            )
            if tkt:
                c['open_tickets'] = tkt['cnt']
        except:
            pass
    
    return render_template('odp/map.html', odps=odps, customers=customers)

@odp_bp.route('/topology')
def topology():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    # Ambil data hirarki infrastruktur
    olts = execute_query("SELECT id, name, ip_address, brand, total_ports FROM olts", fetch=True) or []
    odcs = execute_query("SELECT id, name, capacity, address, olt_id FROM odcs", fetch=True) or []
    odps = execute_query("SELECT id, name, odc_id, capacity FROM odps", fetch=True) or []
    
    return render_template('odp/topology.html', olts=olts, odcs=odcs, odps=odps)


@odp_bp.route('/add', methods=['GET', 'POST'])
def add():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        name = request.form.get('name')
        address = request.form.get('address')
        coordinates = request.form.get('coordinates')
        capacity = request.form.get('capacity')
        odc_id = request.form.get('odc_id')
        
        try:
            execute_query(
                "INSERT INTO odps (name, address, coordinates, capacity, odc_id) VALUES (%s, %s, %s, %s, %s)",
                (name, address, coordinates, capacity, odc_id if odc_id else None)
            )
            flash('ODP berhasil ditambahkan.', 'success')
            return redirect(url_for('odp.index'))
        except Exception as e:
            flash(f'Gagal menambah ODP: {e}', 'error')
            
    odcs = execute_query("SELECT id, name FROM odcs ORDER BY name", fetch=True) or []
    return render_template('odp/form.html', odcs=odcs)

@odp_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    odp = execute_query("SELECT * FROM odps WHERE id=%s", (id,), fetch_one=True)
    if not odp:
        return redirect(url_for('odp.index'))
        
    if request.method == 'POST':
        name = request.form.get('name')
        address = request.form.get('address')
        coordinates = request.form.get('coordinates')
        capacity = request.form.get('capacity')
        odc_id = request.form.get('odc_id')
        
        execute_query(
            "UPDATE odps SET name=%s, address=%s, coordinates=%s, capacity=%s, odc_id=%s WHERE id=%s",
            (name, address, coordinates, capacity, odc_id if odc_id else None, id)
        )
        flash('ODP diperbarui.', 'success')
        return redirect(url_for('odp.index'))
        
    odcs = execute_query("SELECT id, name FROM odcs ORDER BY name", fetch=True) or []
    return render_template('odp/form.html', odp=odp, odcs=odcs)

@odp_bp.route('/delete/<int:id>')
def delete(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    # Check if used
    count = execute_query("SELECT COUNT(*) as c FROM customers WHERE odp_id=%s", (id,), fetch_one=True)['c']
    if count > 0:
        flash(f'Gagal hapus: ODP ini sedang digunakan oleh {count} pelanggan.', 'error')
    else:
        execute_query("DELETE FROM odps WHERE id=%s", (id,))
        flash('ODP dihapus.', 'success')
        
    return redirect(url_for('odp.index'))
@odp_bp.route('/update_customer_coords', methods=['POST'])
def update_customer_coords():
    if not session.get('logged_in'):
        return {"success": False, "message": "Unauthorized"}, 401
    
    data = request.json
    cust_id = data.get('id')
    coords = data.get('coordinates')
    
    if cust_id and coords:
        try:
            execute_query("UPDATE customers SET coordinates=%s WHERE id=%s", (coords, cust_id))
            return {"success": True, "message": "Lokasi pelanggan diperbarui."}
        except Exception as e:
            return {"success": False, "message": str(e)}, 500
            
    return {"success": False, "message": "Data tidak lengkap."}, 400

@odp_bp.route('/update_odp_coords', methods=['POST'])
def update_odp_coords():
    if not session.get('logged_in'):
        return {"success": False, "message": "Unauthorized"}, 401
    
    data = request.json
    odp_id = data.get('id')
    coords = data.get('coordinates')
    
    if odp_id and coords:
        try:
            execute_query("UPDATE odps SET coordinates=%s WHERE id=%s", (coords, odp_id))
            return {"success": True, "message": "Lokasi ODP diperbarui."}
        except Exception as e:
            return {"success": False, "message": str(e)}, 500
            
    return {"success": False, "message": "Data tidak lengkap."}, 400

@odp_bp.route('/import', methods=['GET', 'POST'])
def import_excel():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)
            
        if file and file.filename.endswith('.xlsx'):
            import os
            from openpyxl import load_workbook
            from web.app import UPLOAD_FOLDER
            
            filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filepath)
            
            try:
                wb = load_workbook(filepath)
                sheet = wb.active
                
                count = 0
                # Assuming Header: Name, Address, Coordinates, Capacity, ODC Name
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, address, coords, capacity, odc_name = row
                    if not name: continue
                    
                    odc_id = None
                    if odc_name:
                        odc = execute_query("SELECT id FROM odcs WHERE name=%s", (odc_name,), fetch_one=True)
                        if odc: odc_id = odc['id']

                    execute_query(
                        "INSERT INTO odps (name, address, coordinates, capacity, odc_id) VALUES (%s, %s, %s, %s, %s) "
                        "ON DUPLICATE KEY UPDATE address=%s, coordinates=%s, capacity=%s, odc_id=%s",
                        (name, address or '', coords or '', capacity or 8, odc_id,
                         address or '', coords or '', capacity or 8, odc_id)
                    )
                    count += 1
                
                flash(f'Berhasil mengimport {count} ODP.', 'success')
                return redirect(url_for('odp.index'))
            except Exception as e:
                flash(f'Error reading excel: {e}', 'error')
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    
    return render_template('odp/import.html')
