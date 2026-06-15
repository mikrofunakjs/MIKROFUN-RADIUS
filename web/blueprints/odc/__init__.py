from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify
from web.database import execute_query
from web.decorators import admin_required

odc_bp = Blueprint('odc', __name__)

@odc_bp.route('/')
@admin_required
def index():
    q = request.args.get('q', '')
    query = (
        "SELECT o.*, (SELECT COUNT(*) FROM odps WHERE odc_id = o.id) as odp_count "
        "FROM odcs o"
    )
    params = []
    if q:
        query += " WHERE o.name LIKE %s OR o.address LIKE %s"
        params.extend([f"%{q}%", f"%{q}%"])
    query += " ORDER BY o.name"
    odcs = execute_query(query, tuple(params), fetch=True) or []
    return render_template('odc/list.html', odcs=odcs)

@odc_bp.route('/add', methods=['GET', 'POST'])
@admin_required
def add():
    if request.method == 'POST':
        name = request.form.get('name')
        address = request.form.get('address')
        coords = request.form.get('coordinates')
        capacity = request.form.get('capacity', 12)
        
        ok = execute_query(
            "INSERT INTO odcs (name, address, coordinates, capacity) VALUES (%s, %s, %s, %s)",
            (name, address, coords, capacity)
        )
        if ok:
            flash('ODC berhasil ditambahkan', 'success')
            return redirect(url_for('odc.index'))
        else:
            flash('Gagal menambahkan ODC', 'error')
            
    return render_template('odc/add.html')

@odc_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit(id):
    odc = execute_query("SELECT * FROM odcs WHERE id=%s", (id,), fetch_one=True)
    if not odc:
        flash('ODC tidak ditemukan', 'error')
        return redirect(url_for('odc.index'))
        
    if request.method == 'POST':
        name = request.form.get('name')
        address = request.form.get('address')
        coords = request.form.get('coordinates')
        capacity = request.form.get('capacity')
        
        ok = execute_query(
            "UPDATE odcs SET name=%s, address=%s, coordinates=%s, capacity=%s WHERE id=%s",
            (name, address, coords, capacity, id)
        )
        if ok:
            flash('ODC berhasil diupdate', 'success')
            return redirect(url_for('odc.index'))
        else:
            flash('Gagal update ODC', 'error')
            
    return render_template('odc/edit.html', odc=odc)

@odc_bp.route('/delete/<int:id>')
@admin_required
def delete(id):
    execute_query("UPDATE odps SET odc_id = NULL WHERE odc_id = %s", (id,))
    execute_query("DELETE FROM odcs WHERE id=%s", (id,))
    flash('ODC berhasil dihapus', 'success')
    return redirect(url_for('odc.index'))

@odc_bp.route('/import', methods=['GET', 'POST'])
@admin_required
def import_excel():
    if request.method == 'GET':
        return render_template('odc/import.html')
        
    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('File harus .xlsx', 'error')
        return redirect(url_for('odc.import_excel'))
        
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        
        success = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            data = dict(zip(headers, row))
            execute_query(
                "INSERT IGNORE INTO odcs (name, address, coordinates, capacity) VALUES (%s, %s, %s, %s)",
                (data.get('name'), data.get('address'), data.get('coordinates'), data.get('capacity', 12))
            )
            success += 1
        flash(f'Berhasil import {success} ODC', 'success')
    except Exception as e:
        flash(f'Error import: {e}', 'error')
        
    return redirect(url_for('odc.index'))
