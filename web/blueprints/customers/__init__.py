"""Customers Blueprint - PPPoE User Management"""
from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify
from web.database import execute_query
from web.decorators import cs_or_admin_required
from werkzeug.security import generate_password_hash
import datetime
import math

customers_bp = Blueprint('customers', __name__)

@customers_bp.route('/', methods=['GET'])
@cs_or_admin_required
def index():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    q = request.args.get('q', '')
    status_filter = request.args.get('status', '')
    profile_id = request.args.get('profile_id', '')
    page = max(1, int(request.args.get('page', 1)))
    per_page = 25
    offset = (page - 1) * per_page
    
    where_clause = "WHERE (c.mac_address IS NULL OR c.mac_address = '')"
    params = []
    
    if q:
        where_clause += " AND (c.name LIKE %s OR c.username LIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])
        
    if status_filter:
        where_clause += " AND c.status = %s"
        params.append(status_filter)
        
    if profile_id:
        where_clause += " AND c.profile_id = %s"
        params.append(profile_id)

    # 1. Count Total for Pagination
    total_count = execute_query(f"SELECT COUNT(*) as c FROM customers c {where_clause}", tuple(params), fetch_one=True)['c']
    total_pages = math.ceil(total_count / per_page)
    
    # 2. Optimized Main Query with LIMIT
    query = (
        "SELECT c.*, p.name as profile_name, p.price as profile_price, p.tax_percent, r.name as router_name, "
        "(SELECT COUNT(*) FROM active_sessions s WHERE LOWER(TRIM(s.username)) = LOWER(TRIM(c.username))) as active_session_count, "
        "(SELECT COALESCE(SUM(acctinputoctets),0) FROM radacct ra WHERE LOWER(TRIM(ra.username)) = LOWER(TRIM(c.username))) as total_upload, "
        "(SELECT COALESCE(SUM(acctoutputoctets),0) FROM radacct ra WHERE LOWER(TRIM(ra.username)) = LOWER(TRIM(c.username))) as total_download "
        "FROM customers c "
        "LEFT JOIN profiles p ON c.profile_id = p.id "
        "LEFT JOIN routers r ON c.router_id = r.id "
        f"{where_clause} "
        "ORDER BY c.created_at DESC LIMIT %s OFFSET %s"
    )
    
    p_params = list(params)
    p_params.extend([per_page, offset])
    customers = execute_query(query, tuple(p_params), fetch=True)
    
    # Fetch profiles for the dropdown filter
    profiles = execute_query("SELECT id, name FROM profiles WHERE type='pppoe' ORDER BY name", fetch=True)
    
    # ADVANCED STATS CALCULATION (Keep this efficient)
    stats_query = (
        "SELECT "
        "  COUNT(*) as total, "
        "  SUM(CASE WHEN ra.acctstoptime IS NULL AND ra.username IS NOT NULL THEN 1 ELSE 0 END) as online, "
        "  SUM(CASE WHEN ra.acctstoptime IS NOT NULL OR ra.username IS NULL THEN 1 ELSE 0 END) as offline, "
        "  SUM(CASE WHEN c.status = 'isolir' THEN 1 ELSE 0 END) as isolir, "
        "  SUM(CASE WHEN c.status = 'active' THEN COALESCE(p.price, 0) ELSE 0 END) as omzet "
        "FROM customers c "
        "LEFT JOIN profiles p ON c.profile_id = p.id "
        "LEFT JOIN (SELECT username, acctstoptime FROM radacct ORDER BY radacctid DESC LIMIT 1) ra ON c.username = ra.username "
        "WHERE (c.mac_address IS NULL OR c.mac_address = '') "
    )
    stats = execute_query(stats_query, fetch_one=True)
    
    # Fetch routers for API import modal
    routers = execute_query("SELECT id, name, vpn_ip FROM routers", fetch=True)
    
    return render_template('customers/list.html', 
                           customers=customers or [], 
                           stats=stats,
                           routers=routers or [],
                           profiles=profiles or [],
                           page=page,
                           total_pages=total_pages,
                           q=q,
                           status_f=status_filter,
                           profile_f=profile_id)

@customers_bp.route('/send_bill/<int:id>', methods=['POST'])
@cs_or_admin_required
def send_bill(id):
    if not session.get('logged_in'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    customer = execute_query(
        "SELECT c.*, p.price as profile_price FROM customers c LEFT JOIN profiles p ON c.profile_id = p.id WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    if not customer:
        return jsonify({'success': False, 'message': 'Customer not found.'}), 404

    phone = customer.get('phone')
    if not phone:
        return jsonify({'success': False, 'message': 'User tidak memiliki nomor referensi WA.'}), 400

    # Safely handle billing properties
    try:
        raw_amount = customer.get('profile_price')
        if raw_amount is None:
            raw_amount = 0
            
        amount = '{:,.0f}'.format(raw_amount).replace(',', '.')
        due = customer.get('due_date') or '-'
        status = customer.get('status', 'aktif').upper()
        
        msg = (f"*INFO TAGIHAN MIKROFUN*\n\n"
               f"Yth. Bapak/Ibu *{customer.get('name')}*,\n\n"
               f"Kami informasikan rincian tagihan internet PPPoE Anda:\n"
               f"🔹 *Username:* {customer.get('username')}\n"
               f"🔹 *Tagihan:* Rp {amount}\n"
               f"🔹 *Jatuh Tempo:* {due}\n"
               f"🔹 *Status:* {status}\n\n"
               f"Mohon segera melakukan pembayaran agar layanan tetap aktif. Terima kasih.")

        from web.wa_helper import send_wa
        if send_wa(phone, msg):
            return jsonify({'success': True, 'message': 'Tagihan berhasil dikirim ke antrean.'})
        else:
            return jsonify({'success': False, 'message': 'Gagal mengirim WA.'}), 500
    except Exception as e:
        print(f"send_bill error for customer {id}: {e}")
        return jsonify({'success': False, 'message': 'Gagal mengirim tagihan. Silakan coba lagi.'}), 500

@customers_bp.route('/bulk_action', methods=['POST'])
def bulk_action():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    action = request.form.get('action')
    customer_ids = request.form.getlist('customer_ids')
    
    if not customer_ids:
        flash('Tidak ada customer yang dipilih!', 'warning')
        return redirect(url_for('customers.index'))
        
    success_count = 0

    if action == 'check_status':
        # Optimized Bulk Check: Group by Router
        format_strings = ','.join(['%s'] * len(customer_ids))
        query = (
             f"SELECT c.username, r.id as router_id, r.vpn_ip, r.api_user, r.api_password, r.api_port "
             f"FROM customers c LEFT JOIN routers r ON c.router_id = r.id "
             f"WHERE c.id IN ({format_strings})"
        )
        customers_to_check = execute_query(query, tuple(customer_ids), fetch=True)
        
        if customers_to_check:
            from web.mikrotik_api import MikrotikApi
            
            # Group by Router
            router_groups = {}
            for c in customers_to_check:
                 if not c['router_id']: continue 
                 if c['router_id'] not in router_groups:
                     router_groups[c['router_id']] = {'creds': c, 'users': []}
                 router_groups[c['router_id']]['users'].append(c['username'])
            
            # Process Each Router
            for rid, group in router_groups.items():
                 creds = group['creds']
                 target_users = set(group['users'])
                 
                 try:
                     api = MikrotikApi(creds['vpn_ip'], creds.get('api_port', 8728))
                     if api.connect() and api.login(creds['api_user'], creds['api_password']):
                         # Fetch ALL active users once
                         active_rows = api.query(['/ppp/active/print'])
                         api.close()
                         
                         active_map = {row.get('name'): row for row in active_rows if row.get('name')}
                         
                         for user in target_users:
                             if user in active_map:
                                 # ONLINE
                                 pkt = active_map[user]
                                 nas_ip = creds['vpn_ip']
                                 session_id = pkt.get('.id', 'unknown')
                                 execute_query(
                                     "INSERT INTO active_sessions (username, nas_ip, acct_session_id) VALUES (%s, %s, %s) "
                                     "ON DUPLICATE KEY UPDATE updated_at=NOW()",
                                     (user, nas_ip, session_id)
                                 )
                             else:
                                 # OFFLINE
                                 execute_query("DELETE FROM active_sessions WHERE username=%s", (user,))
                         
                         success_count += len(target_users)
                 except Exception as e:
                     print(f"Bulk Check Error {creds['vpn_ip']}: {e}")

    else:
        # Standard Item-by-Item Loop
        for cid in customer_ids:
            try:
                if action == 'delete':
                    # Get info for API kick
                    customer = execute_query(
                        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
                        "FROM customers c LEFT JOIN routers r ON c.router_id = r.id WHERE c.id=%s", 
                        (cid,), fetch_one=True
                    )
                    if customer and customer['api_user']:
                        try_api_disconnect(customer, customer['username'])
                    
                    execute_query("DELETE FROM customers WHERE id=%s", (cid,))
                    
                elif action == 'isolir':
                    customer = execute_query(
                        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
                        "FROM customers c LEFT JOIN routers r ON c.router_id = r.id WHERE c.id=%s", 
                        (cid,), fetch_one=True
                    )
                    if customer and customer['api_user']:
                        try_api_disconnect(customer, customer['username'])
                    
                    # Force remove session
                    execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))
                        
                    execute_query("UPDATE customers SET status='isolir' WHERE id=%s", (cid,))
                    
                elif action == 'activate':
                    execute_query("UPDATE customers SET status='active' WHERE id=%s", (cid,))
                    
                success_count += 1
            except Exception as e:
                continue
            
    flash(f'Berhasil memproses {success_count} customer.', 'success')
    return redirect(url_for('customers.index'))

@customers_bp.route('/add', methods=['GET', 'POST'])
@cs_or_admin_required
def add():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    if request.method == 'POST':
        name = request.form.get('name')
        username = request.form.get('username')
        password = request.form.get('password')
        service_type = request.form.get('service_type', 'pppoe')
        profile_id = request.form.get('profile_id')
        router_id = request.form.get('router_id')
        phone = request.form.get('phone', '')
        address = request.form.get('address', '')
        due_date = request.form.get('due_date') or None
        billing_type = request.form.get('billing_type', 'prepaid')

        odp_id = request.form.get('odp_id') or None
        port_number = request.form.get('port_number') or None
        coordinates = request.form.get('coordinates', '')
        
        mac_address = request.form.get('mac_address', '').strip().upper() or None
        static_ip = request.form.get('static_ip', '').strip() or None

        # CHECK LIMIT FOR FREE USERS
        from web.license_service import is_premium
        if not is_premium():
             current_count = execute_query("SELECT COUNT(*) as c FROM customers", fetch_one=True)['c']
             if current_count >= 100:
                  flash('Batas Free: 100 pelanggan. Upgrade ke Premium untuk unlimited.', 'error')
                  return redirect(url_for('customers.add'))

        # Hash password before storing
        hashed_password = generate_password_hash(password)

        result = execute_query(
            "INSERT INTO customers (name, username, password, service_type, profile_id, router_id, phone, address, due_date, status, odp_id, port_number, coordinates, mac_address, static_ip, billing_type) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'active',%s,%s,%s,%s,%s,%s)",
            (name, username, hashed_password, service_type, profile_id, router_id, phone, address, due_date, odp_id, port_number, coordinates, mac_address, static_ip, billing_type)
        )
        if result:
            from web.telegram_helper import send_telegram_message
            send_telegram_message(f"👤 *Pelanggan Baru (PPPoE)*\n\nNama: {name}\nUsername: {username}\nAlamat: {address}")
            
            from web.blueprints.notifications import add_notification
            add_notification(
                title="Customer Created",
                message=f"Created customer {name} ({username})",
                category="success"
            )
            flash('Customer berhasil dibuat', 'success')
        else:
            flash('Gagal membuat customer', 'error')
            
        return redirect(url_for('customers.index'))
        
    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []
    profiles = execute_query("SELECT * FROM profiles WHERE type=%s ORDER BY name", ('pppoe',), fetch=True) or []
    odps = execute_query("SELECT * FROM odps ORDER BY name", fetch=True) or []
    
    return render_template('customers/add.html', routers=routers, profiles=profiles, odps=odps)

@customers_bp.route('/check_status/<int:id>')
def check_status(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c LEFT JOIN routers r ON c.router_id = r.id WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    
    if not customer:
        flash('Customer tidak ditemukan', 'error')
        return redirect(url_for('customers.index'))
        
    username = customer['username']
    
    # Logic: Connect -> Check. If fail connect/check -> Offline
    try:
        from web.mikrotik_api import MikrotikApi
        
        api = MikrotikApi(customer['vpn_ip'], customer.get('api_port', 8728))
        if api.connect():
             if api.login(customer['api_user'], customer['api_password']):
                 user_data = api.get_active_user(username)
                 api.close()
                 
                 if user_data:
                     # ONLINE
                     nas_ip = customer['vpn_ip']
                     session_id = user_data.get('.id', 'unknown')
                     execute_query(
                         "INSERT INTO active_sessions (username, nas_ip, acct_session_id) VALUES (%s, %s, %s) "
                         "ON DUPLICATE KEY UPDATE updated_at=NOW()",
                         (username, nas_ip, session_id)
                     )
                     flash(f'User {username} is ONLINE', 'success')
                 else:
                     # OFFLINE (User not found in active list)
                     execute_query("DELETE FROM active_sessions WHERE username=%s", (username,))
                     flash(f'User {username} is OFFLINE', 'warning')
             else:
                 api.close()
                 flash('Gagal login ke Mikrotik', 'error')
        else:
             # Connect Fail -> Assume Offline
             execute_query("DELETE FROM active_sessions WHERE username=%s", (username,))
             flash(f'Gagal koneksi ke Mikrotik ({customer["vpn_ip"]}). Status set OFFLINE.', 'error')
             
    except Exception as e:
        flash(f'Error Check Status: {e}', 'error')
        
    return redirect(url_for('customers.index'))

@customers_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@cs_or_admin_required
def edit(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    customer = execute_query("SELECT * FROM customers WHERE id=%s", (id,), fetch_one=True)
    if not customer:
        flash('Customer tidak ditemukan!', 'error')
        return redirect(url_for('customers.index'))

    if request.method == 'POST':
        due_date = request.form.get('due_date') or None
        mac_address = request.form.get('mac_address', '').strip().upper() or None
        static_ip = request.form.get('static_ip', '').strip() or None
        billing_type = request.form.get('billing_type', 'prepaid')
        new_password = request.form.get('password', '').strip()
        
        if new_password:
            # Hash new password if provided
            hashed_password = generate_password_hash(new_password)
            execute_query(
                "UPDATE customers SET name=%s, password=%s, profile_id=%s, router_id=%s, "
                "phone=%s, address=%s, due_date=%s, odp_id=%s, port_number=%s, coordinates=%s, mac_address=%s, static_ip=%s, billing_type=%s WHERE id=%s",
                (request.form.get('name'), hashed_password,
                 request.form.get('profile_id'), request.form.get('router_id'),
                 request.form.get('phone',''), request.form.get('address',''), due_date,
                 request.form.get('odp_id') or None, request.form.get('port_number') or None, 
                 request.form.get('coordinates',''), mac_address, static_ip, billing_type, id)
            )
        else:
            # No password change — don't touch password column
            execute_query(
                "UPDATE customers SET name=%s, profile_id=%s, router_id=%s, "
                "phone=%s, address=%s, due_date=%s, odp_id=%s, port_number=%s, coordinates=%s, mac_address=%s, static_ip=%s, billing_type=%s WHERE id=%s",
                (request.form.get('name'),
                 request.form.get('profile_id'), request.form.get('router_id'),
                 request.form.get('phone',''), request.form.get('address',''), due_date,
                 request.form.get('odp_id') or None, request.form.get('port_number') or None, 
                 request.form.get('coordinates',''), mac_address, static_ip, billing_type, id)
            )
        flash('Customer berhasil diupdate!', 'success')
        return redirect(url_for('customers.index'))

    profiles = execute_query("SELECT * FROM profiles ORDER BY name", fetch=True) or []
    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []
    odps = execute_query("SELECT * FROM odps ORDER BY name", fetch=True) or []
    return render_template('customers/edit.html', customer=customer, profiles=profiles, routers=routers, odps=odps)

@customers_bp.route('/sync_online', methods=['POST'])
@cs_or_admin_required
def sync_online():
    """Sync Online status via Mikrotik API for all routers"""
    routers = execute_query("SELECT * FROM routers WHERE api_user IS NOT NULL AND api_user != ''", fetch=True) or []
    if not routers:
        return jsonify({'success': False, 'message': 'No routers with API configured.'})

    from web.mikrotik_api import MikrotikApi
    synced_count = 0
    errors = []

    # Optional: Clear old sessions first? No, let's just update/insert.
    # To be really accurate, we should mark which router we are syncing.
    
    for r in routers:
        try:
            api = MikrotikApi(r['vpn_ip'], int(r['api_port'] or 8728))
            if api.connect():
                if api.login(r['api_user'], r['api_password']):
                    sessions = api.get_all_active_sessions()
                    if sessions:
                        for s in sessions:
                            uname = s['username'].strip().lower()
                            # Update active_sessions
                            execute_query(
                                "INSERT INTO active_sessions (username, nas_ip, acct_session_id) VALUES (%s, %s, %s) "
                                "ON DUPLICATE KEY UPDATE updated_at=NOW(), nas_ip=%s",
                                (uname, r['vpn_ip'], s['session_id'], r['vpn_ip'])
                            )
                            synced_count += 1
                    api.close()
                else:
                    errors.append(f"Login failed: {r['name']}")
            else:
                errors.append(f"Connect failed: {r['name']}")
        except Exception as e:
            errors.append(f"Error {r['name']}: {str(e)}")

    # Clean up sessions that haven't been updated in 5 minutes (meaning they are no longer in API list)
    execute_query("DELETE FROM active_sessions WHERE updated_at < DATE_SUB(NOW(), INTERVAL 5 MINUTE)")

    return jsonify({
        'success': True, 
        'message': f'Synced {synced_count} sessions from {len(routers)} routers.',
        'errors': errors
    })

def try_api_disconnect(router_data, username):
    """Helper to disconnect user via Mikrotik API"""
    if not router_data or not router_data.get('vpn_ip') or not router_data.get('api_user'):
        return False, "Router API not configured"
        
    try:
        from web.mikrotik_api import MikrotikApi
        api = MikrotikApi(router_data['vpn_ip'], int(router_data.get('api_port') or 8728))
        
        if not api.login(router_data['api_user'], router_data['api_password']):
            return False, "API Login Failed"
            
        success, msg = api.kick_user(username)
        api.close()
        return success, msg
    except Exception as e:
        return False, str(e)

@customers_bp.route('/delete/<int:id>')
def delete(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    # Get user info and router info BEFORE deleting
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    
    msg = 'Customer berhasil dihapus!'
    
    # 1. API Kick
    if customer and customer['api_user']:
        success, reason = try_api_disconnect(customer, customer['username'])
        if success:
            msg += ' Session diputus (API).'
        else:
            msg += f' (API Kick Warning: {reason})'
    
    # Force remove session
    execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))

    # 2. Delete from DB
    execute_query("DELETE FROM customers WHERE id=%s", (id,))
    
    flash(msg, 'success')
    return redirect(url_for('customers.index'))

@customers_bp.route('/isolir/<int:id>')
def isolir(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    # Get user info
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE c.id=%s", 
        (id,), fetch_one=True
    )
    
    # 1. Update status in DB
    execute_query("UPDATE customers SET status='isolir' WHERE id=%s", (id,))
    
    msg = 'Customer diisolir!'
    
    # 2. API Kick
    if customer and customer['api_user']:
        success, reason = try_api_disconnect(customer, customer['username'])
        if success:
            msg += ' Session diputus (API).'
        else:
            msg += f' (API Kick Warning: {reason})'
    else:
        msg += ' (Note: API belum disetting di router, user tidak otomatis DC)'
            
    # 3. Force Remove Session from DB (UI Feedback)
    execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))

    flash(msg, 'success')
    return redirect(url_for('customers.index'))

@customers_bp.route('/activate/<int:id>')
def activate(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    execute_query("UPDATE customers SET status='active' WHERE id=%s", (id,))
    flash('Customer diaktifkan!', 'success')
    return redirect(url_for('customers.index'))

@customers_bp.route('/download_template')
@cs_or_admin_required
def download_template():
    """Download blank Excel template for PPPoE bulk import."""
    import io
    from flask import send_file
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        flash('Library openpyxl belum terinstall. Jalankan: pip install openpyxl', 'error')
        return redirect(url_for('customers.index'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PPPoE Customers"

    headers = ['name', 'username', 'password', 'profile_name', 'phone', 'address', 'due_date', 'router_name']
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Example row
    ws.append(['Budi Santoso', 'budi001', 'pass001', 'Paket 10Mbps', '08123456789', 'Jl. Mawar No. 1', '2025-12-31', 'Router Main'])

    # Widen columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='template_import_pppoe.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@customers_bp.route('/import', methods=['GET', 'POST'])
@cs_or_admin_required
def import_excel():
    """Import PPPoE customers from Excel file."""
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    if request.method == 'GET':
        # Get routers for dropdown info
        routers = execute_query("SELECT id, name, vpn_ip FROM routers WHERE api_user IS NOT NULL AND api_user != ''", fetch=True) or []
        return render_template('customers/import.html', routers=routers)

    # POST - process upload
    try:
        import openpyxl
    except ImportError:
        flash('Library openpyxl belum terinstall. Jalankan: pip install openpyxl', 'error')
        return redirect(url_for('customers.index'))

    file = request.files.get('excel_file')
    if not file or not file.filename.endswith('.xlsx'):
        flash('File harus berformat .xlsx (Excel)', 'error')
        return redirect(url_for('customers.import_excel'))

    skip_duplicates = request.form.get('skip_duplicates') == '1'

    results = []
    success_count = 0
    error_count = 0

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Get headers from row 1
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

        required = {'name', 'username', 'password', 'profile_name'}
        if not required.issubset(set(headers)):
            flash(f'Kolom wajib tidak ditemukan. Kolom yang diperlukan: {", ".join(required)}', 'error')
            return redirect(url_for('customers.import_excel'))

        # Get all profiles for lookup
        profiles = execute_query("SELECT id, name FROM profiles", fetch=True) or []
        profile_map = {p['name'].lower(): p['id'] for p in profiles}

        # Get all ODPs for lookup
        odps = execute_query("SELECT id, name FROM odps", fetch=True) or []
        odp_map = {o['name'].lower(): o['id'] for o in odps}

        # Get all routers for lookup
        routers = execute_query("SELECT id, name, vpn_ip FROM routers", fetch=True) or []
        router_map = {r['name'].lower(): r['id'] for r in routers}

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {}
            try:
                if not any(row):
                    continue  # skip empty rows

                # Secure dictionary creation with header check
                row_data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row) if i < len(headers)}

                name     = row_data.get('name', '')
                username = row_data.get('username', '')
                password = row_data.get('password', '')
                profile_name = row_data.get('profile_name', '')
                phone    = row_data.get('phone', '')
                address  = row_data.get('address', '')
                due_date = row_data.get('due_date', '') or None
                odp_name = row_data.get('odp_name', '')
                router_name = row_data.get('router_name', '')

                # Validate required
                if not name or not username or not password or not profile_name:
                    results.append({'row': row_num, 'name': name or username, 'status': 'error',
                                    'message': 'Kolom name, username, password, profile_name wajib diisi'})
                    error_count += 1
                    continue

                # Lookup profile
                profile_id = profile_map.get(profile_name.lower())
                if not profile_id:
                    results.append({'row': row_num, 'name': name, 'status': 'error',
                                    'message': f'Profile "{profile_name}" tidak ditemukan di database'})
                    error_count += 1
                    continue

                # Resolve ODP if provided
                odp_id = None
                if odp_name:
                    odp_id = odp_map.get(odp_name.lower())
                    if not odp_id:
                        results.append({'row': row_num, 'name': name, 'status': 'error',
                                        'message': f'ODP "{odp_name}" tidak ditemukan di database'})
                        error_count += 1
                        continue

                # Resolve Router if provided
                router_id = None
                if router_name:
                    router_id = router_map.get(router_name.lower())
                    if not router_id:
                        results.append({'row': row_num, 'name': name, 'status': 'error',
                                        'message': f'Router "{router_name}" tidak ditemukan di database'})
                        error_count += 1
                        continue

                # Check duplicate
                existing = execute_query("SELECT id FROM customers WHERE username=%s", (username,), fetch_one=True)
                if existing:
                    if skip_duplicates:
                        results.append({'row': row_num, 'name': name, 'status': 'skip',
                                        'message': f'Username "{username}" sudah ada (dilewati)'})
                        continue
                    else:
                        results.append({'row': row_num, 'name': name, 'status': 'error',
                                        'message': f'Username "{username}" sudah ada'})
                        error_count += 1
                        continue

                # Insert
                hashed_pw = generate_password_hash(password)
                ok = execute_query(
                    "INSERT INTO customers (name, username, password, service_type, profile_id, phone, address, due_date, status, billing_type, odp_id, router_id) "
                    "VALUES (%s,%s,%s,'pppoe',%s,%s,%s,%s,'active', %s, %s, %s)",
                    (name, username, hashed_pw, profile_id, phone, address, due_date or None, row_data.get('billing_type', 'prepaid'), odp_id, router_id)
                )
                if ok:
                    results.append({'row': row_num, 'name': name, 'status': 'success', 'message': 'Berhasil ditambahkan'})
                    success_count += 1
                else:
                    results.append({'row': row_num, 'name': name, 'status': 'error', 'message': 'Gagal menyimpan ke database'})
                    error_count += 1
            except Exception as e:
                results.append({'row': row_num, 'name': row_data.get('name', 'Unknown'), 'status': 'error', 'message': f'System Error: {str(e)}'})
                error_count += 1

    except Exception as e:
        flash(f'Gagal memproses file Excel: {e}', 'error')
        return redirect(url_for('customers.import_excel'))

    return render_template('customers/import_result.html',
                           results=results,
                           success_count=success_count,
                           error_count=error_count)

@customers_bp.route('/import_api_mikrotik', methods=['POST'])
@cs_or_admin_required
def import_api_mikrotik():
    router_id = request.form.get('router_id')
    profile_id = request.form.get('profile_id')
    billing_type = request.form.get('billing_type')
    due_date = request.form.get('due_date')
    
    if not all([router_id, profile_id, billing_type, due_date]):
        flash('Semua field harus diisi!', 'error')
        return redirect(url_for('customers.index'))
        
    router = execute_query("SELECT * FROM routers WHERE id=%s", (router_id,), fetch_one=True)
    if not router:
        flash('Router tidak ditemukan.', 'error')
        return redirect(url_for('customers.index'))
        
    profile = execute_query("SELECT * FROM profiles WHERE id=%s", (profile_id,), fetch_one=True)
    if not profile:
        flash('Profil / Paket tidak ditemukan.', 'error')
        return redirect(url_for('customers.index'))
        
    # Connect to router
    from web.mikrotik_api import MikrotikApi
    # Router table uses: vpn_ip, api_port, api_user, api_password
    api = MikrotikApi(router['vpn_ip'], router['api_port'])
    
    try:
        if not api.connect():
            flash(f"Gagal koneksi ke port API {router['api_port']} di {router['vpn_ip']}", 'error')
            return redirect(url_for('customers.index'))
            
        if not api.login(router['api_user'], router['api_password']):
            flash("Login API Gagal: Username atau Password Router salah.", 'error')
            return redirect(url_for('customers.index'))
    except Exception as e:
        flash(f'Gagal terhubung ke router: {e}', 'error')
        return redirect(url_for('customers.index'))
        
    try:
        # Use query() instead of get_resources()
        secrets = api.query(['/ppp/secret/print'])
        success_count = 0
        skip_count = 0
        
        for secret in secrets:
            username = secret.get('name')
            password = secret.get('password', '')
            
            if not username:
                continue
                
            # Skip if username already exists in MicroFun customers table
            existing = execute_query("SELECT id FROM customers WHERE username=%s", (username,), fetch_one=True)
            if existing:
                skip_count += 1
                continue
                
            # Prepare data
            name = username 
            phone = '12345678'
            address = 'client'
            
            # Insert to customers
            hashed_pw = generate_password_hash(password)
            ok = execute_query(
                "INSERT INTO customers (name, username, password, service_type, profile_id, router_id, phone, address, due_date, status, billing_type, mac_address, created_at) "
                "VALUES (%s,%s,%s,'pppoe',%s,%s,%s,%s,%s,'active',%s,NULL,NOW())",
                (name, username, hashed_pw, profile_id, router_id, phone, address, due_date, billing_type)
            )
            
            if ok:
                # Insert to radcheck
                execute_query(
                    "INSERT INTO radcheck (username, attribute, op, value) VALUES (%s, 'Cleartext-Password', ':=', %s)",
                    (username, password)
                )
                
                # Insert to radusergroup
                execute_query(
                    "INSERT INTO radusergroup (username, groupname, priority) VALUES (%s, %s, 1)",
                    (username, profile['name'])
                )
                
                success_count += 1
            else:
                skip_count += 1
            
        if success_count > 0:
            flash(f'Berhasil mengimport {success_count} data PPPoE baru. ({skip_count} dilewati/duplikat)', 'success')
        else:
            flash(f'Tidak ada data baru yang berhasil diimport. ({skip_count} duplikat / kemungkinan username sudah ada di sistem)', 'warning')
        
    except Exception as e:
        flash(f'Gagal memproses data MikroTik: {e}', 'error')
        
    finally:
        api.close()
        
    return redirect(url_for('customers.index'))

@customers_bp.route('/active', methods=['GET'])
@cs_or_admin_required
def active():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    import datetime
    today = datetime.date.today()
    
    query = (
        "SELECT c.*, p.name as profile_name, r.name as router_name, "
        "MAX(s.nas_ip) as session_nas_ip, MAX(s.acct_session_id) as acct_session_id "
        "FROM customers c "
        "LEFT JOIN profiles p ON c.profile_id = p.id "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "LEFT JOIN active_sessions s ON LOWER(TRIM(c.username)) = LOWER(TRIM(s.username)) "
        "WHERE c.status = 'active' AND (c.mac_address IS NULL OR c.mac_address = '') "
        "GROUP BY c.id, p.name, r.name "
        "ORDER BY c.due_date ASC"
    )
    
    customers = execute_query(query, fetch=True) or []
    
    enriched = []
    for c in customers:
        days_left = None
        if c.get('due_date'):
            delta = c['due_date'] - today
            days_left = delta.days
            
        enriched.append({
            **c,
            'days_left': days_left,
            'is_online': True if c.get('acct_session_id') else False
        })
        
    return render_template('customers/active.html', customers=enriched)

@customers_bp.route('/isolir-list', methods=['GET'])
@cs_or_admin_required
def isolir_list():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
    
    query = (
        "SELECT c.*, p.name as profile_name, r.name as router_name "
        "FROM customers c "
        "LEFT JOIN profiles p ON c.profile_id = p.id "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE c.status = 'isolir' AND (c.mac_address IS NULL OR c.mac_address = '') "
        "ORDER BY c.updated_at DESC"
    )
    
    customers = execute_query(query, fetch=True) or []
    return render_template('customers/isolir.html', customers=customers)

@customers_bp.route('/kick/<int:id>', methods=['POST'])
@cs_or_admin_required
def kick_session(id):
    customer = execute_query(
        "SELECT c.username, r.vpn_ip, r.api_user, r.api_password, r.api_port "
        "FROM customers c "
        "LEFT JOIN routers r ON c.router_id = r.id "
        "WHERE c.id=%s", (id,), fetch_one=True
    )
    
    if not customer:
        return jsonify({'success': False, 'message': 'Pelanggan tidak ditemukan.'}), 404
        
    if not customer['vpn_ip'] or not customer['api_user']:
        return jsonify({'success': False, 'message': 'API Router belum disetting.'}), 400

    from web.mikrotik_api import MikrotikApi
    try:
        api = MikrotikApi(customer['vpn_ip'], int(customer['api_port'] or 8728))
        if api.connect() and api.login(customer['api_user'], customer['api_password']):
            # Kick logic
            api.query(['/ppp/active/remove', '=.id=' + customer['username']]) # MikroTik usually identifies by name or ID
            # In Mikrotik API, remove usually needs the .id from print. 
            # A more robust way is querying the ID first.
            active_row = api.get_active_user(customer['username'])
            if active_row:
                api.query(['/ppp/active/remove', '=.id=' + active_row['.id']])
                api.close()
                execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))
                return jsonify({'success': True, 'message': f'Session {customer["username"]} berhasil diputus.'})
            else:
                api.close()
                execute_query("DELETE FROM active_sessions WHERE username=%s", (customer['username'],))
                return jsonify({'success': False, 'message': 'User sedang tidak online di MikroTik.'})
        else:
            return jsonify({'success': False, 'message': 'Gagal login ke MikroTik.'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
