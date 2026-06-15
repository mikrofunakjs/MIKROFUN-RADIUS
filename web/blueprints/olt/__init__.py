from flask import Blueprint, render_template, request, session, redirect, url_for, flash
from web.database import execute_query
from web.decorators import admin_required

olt_bp = Blueprint('olt', __name__)

@olt_bp.route('/')
def index():
    if not session.get('logged_in'): return redirect(url_for('auth.login'))
    olts = execute_query("SELECT * FROM olts ORDER BY name ASC", fetch=True) or []
    return render_template('olt/list.html', olts=olts)

@olt_bp.route('/add', methods=['GET', 'POST'])
def add():
    if not session.get('logged_in'): return redirect(url_for('auth.login'))
    if request.method == 'POST':
        name = request.form.get('name')
        brand = request.form.get('brand')
        ip = request.form.get('ip_address')
        port = request.form.get('api_port', 8728)
        user = request.form.get('username')
        pw = request.form.get('password')
        total = request.form.get('total_ports', 8)
        
        execute_query(
            "INSERT INTO olts (name, brand, ip_address, api_port, username, password, total_ports) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (name, brand, ip, port, user, pw, total)
        )
        flash('OLT berhasil ditambahkan.', 'success')
        return redirect(url_for('olt.index'))
    return render_template('olt/form.html')

@olt_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    if not session.get('logged_in'): return redirect(url_for('auth.login'))
    olt = execute_query("SELECT * FROM olts WHERE id=%s", (id,), fetch_one=True)
    if not olt: return redirect(url_for('olt.index'))
    
    if request.method == 'POST':
        execute_query(
            "UPDATE olts SET name=%s, brand=%s, ip_address=%s, api_port=%s, username=%s, password=%s, total_ports=%s WHERE id=%s",
            (request.form.get('name'), request.form.get('brand'), request.form.get('ip_address'),
             request.form.get('api_port'), request.form.get('username'), request.form.get('password'),
             request.form.get('total_ports'), id)
        )
        flash('OLT diperbarui.', 'success')
        return redirect(url_for('olt.index'))
    return render_template('olt/form.html', olt=olt)

@olt_bp.route('/delete/<int:id>')
def delete(id):
    execute_query("DELETE FROM olts WHERE id=%s", (id,))
    flash('OLT dihapus.', 'success')
    return redirect(url_for('olt.index'))

@olt_bp.route('/import', methods=['GET', 'POST'])
def import_excel():
    if request.method == 'POST':
        file = request.files.get('file')
        if file and file.filename.endswith('.xlsx'):
            import os
            from openpyxl import load_workbook
            from web.app import UPLOAD_FOLDER
            filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filepath)
            try:
                wb = load_workbook(filepath)
                sheet = wb.active
                # Header Mapping
                headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
                
                count = 0
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_data = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row) if i < len(headers)}
                        
                        name = row_data.get('name')
                        if not name: continue
                        
                        brand = row_data.get('brand', '')
                        ip = row_data.get('ip_address', '') or row_data.get('ip', '')
                        user = row_data.get('username', '') or row_data.get('user', '')
                        pw = row_data.get('password', '') or row_data.get('pw', '')
                        total = row_data.get('total_ports', 8) or row_data.get('total', 8)
                        
                        execute_query(
                            "INSERT INTO olts (name, brand, ip_address, username, password, total_ports) VALUES (%s,%s,%s,%s,%s,%s) "
                            "ON DUPLICATE KEY UPDATE brand=%s, ip_address=%s, username=%s, password=%s, total_ports=%s",
                            (name, brand, ip, user, pw, total, brand, ip, user, pw, total)
                        )
                        count += 1
                    except Exception as row_err:
                        print(f"Row {row_idx} Error: {row_err}")
                        continue
                        
                flash(f'Import {count} OLT berhasil.', 'success')
                return redirect(url_for('olt.index'))
            except Exception as e: flash(f'Error: {e}', 'error')
            finally: 
                if os.path.exists(filepath): os.remove(filepath)
    return render_template('olt/import.html')
