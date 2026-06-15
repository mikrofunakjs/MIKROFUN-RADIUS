"""Profiles Blueprint"""
from flask import Blueprint, render_template, request, session, redirect, url_for, flash
from web.database import execute_query

profiles_bp = Blueprint('profiles', __name__)

@profiles_bp.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    type_filter = request.args.get('type', 'pppoe')
    router_filter = request.args.get('router_id', '')

    query = "SELECT p.*, r.name as router_name FROM profiles p LEFT JOIN routers r ON p.router_id = r.id WHERE p.type=%s"
    params = [type_filter]

    if router_filter:
        query += " AND p.router_id = %s"
        params.append(router_filter)

    query += " ORDER BY r.name ASC, p.price ASC"

    profiles = execute_query(query, tuple(params), fetch=True) or []
    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []

    return render_template('profiles/list.html', profiles=profiles, type=type_filter,
                           routers=routers, router_filter=router_filter)


@profiles_bp.route('/add', methods=['GET', 'POST'])
def add():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    type_filter = request.args.get('type', 'pppoe')

    if request.method == 'POST':
        p_type = request.form.get('type', 'pppoe')
        router_id = request.form.get('router_id') or None
        try:
            price = float(request.form.get('price') or 0)
            validity = int(request.form.get('validity') or 0)
            shared_users = int(request.form.get('shared_users') or 1)
            quota_gb = float(request.form.get('quota_limit') or 0)
            quota_limit = int(quota_gb * 1024 * 1024 * 1024) # Convert GB to Bytes
        except:
            price = 0
            validity = 0
            shared_users = 1
            quota_limit = 0

        result = execute_query(
            "INSERT INTO profiles (name, rate_limit, burst_limit, burst_threshold, burst_time, limit_at, pool_name, shared_users, price, validity, validity_unit, quota_limit, description, type, router_id) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (request.form.get('name'), 
             request.form.get('rate_limit'),
             request.form.get('burst_limit') or None,
             request.form.get('burst_threshold') or None,
             request.form.get('burst_time') or None,
             request.form.get('limit_at') or None,
             request.form.get('pool_name', ''), 
             shared_users,
             price,
             validity,
             request.form.get('validity_unit', 'hours'),
             quota_limit,
             request.form.get('description', ''),
             p_type, router_id)
        )
        if result:
            flash(f'Profile {p_type} berhasil ditambahkan!', 'success')
            return redirect(url_for('profiles.index', type=p_type))
        flash('Gagal menambahkan profile!', 'error')

    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []
    return render_template('profiles/add.html', type=type_filter, routers=routers)


@profiles_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    profile = execute_query("SELECT * FROM profiles WHERE id=%s", (id,), fetch_one=True)
    if not profile:
        flash('Profile tidak ditemukan!', 'error')
        return redirect(url_for('profiles.index'))

    current_type = profile.get('type', 'pppoe')

    if request.method == 'POST':
        router_id = request.form.get('router_id') or None
        try:
            price = float(request.form.get('price') or 0)
            validity = int(request.form.get('validity') or 0)
            shared_users = int(request.form.get('shared_users') or 1)
            quota_gb = float(request.form.get('quota_limit') or 0)
            quota_limit = int(quota_gb * 1024 * 1024 * 1024)
        except:
            price = 0
            validity = 0
            shared_users = 1
            quota_limit = 0

        execute_query(
            "UPDATE profiles SET name=%s, rate_limit=%s, burst_limit=%s, burst_threshold=%s, "
            "burst_time=%s, limit_at=%s, pool_name=%s, shared_users=%s, price=%s, validity=%s, validity_unit=%s, quota_limit=%s, description=%s, router_id=%s WHERE id=%s",
            (request.form.get('name'), 
             request.form.get('rate_limit'),
             request.form.get('burst_limit') or None,
             request.form.get('burst_threshold') or None,
             request.form.get('burst_time') or None,
             request.form.get('limit_at') or None,
             request.form.get('pool_name', ''),
             shared_users,
             price,
             validity,
             request.form.get('validity_unit', 'hours'),
             quota_limit,
             request.form.get('description', ''), router_id, id)
        )
        flash('Profile berhasil diupdate!', 'success')
        return redirect(url_for('profiles.index', type=current_type))

    routers = execute_query("SELECT * FROM routers ORDER BY name", fetch=True) or []
    return render_template('profiles/edit.html', profile=profile, routers=routers)


@profiles_bp.route('/delete/<int:id>')
def delete(id):
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))

    profile = execute_query("SELECT * FROM profiles WHERE id=%s", (id,), fetch_one=True)
    redirect_type = profile['type'] if profile else 'pppoe'

    execute_query("DELETE FROM profiles WHERE id=%s", (id,))
    flash('Profile berhasil dihapus!', 'success')
    return redirect(url_for('profiles.index', type=redirect_type))


@profiles_bp.route('/api/by_router')
def api_by_router():
    """JSON: return profiles filtered by router_id for dynamic customer form."""
    router_id = request.args.get('router_id', '')
    p_type = request.args.get('type', 'pppoe')

    if router_id:
        profiles = execute_query(
            "SELECT id, name, price, rate_limit FROM profiles WHERE type=%s AND (router_id=%s OR router_id IS NULL) ORDER BY price ASC",
            (p_type, router_id), fetch=True
        ) or []
    else:
        profiles = execute_query(
            "SELECT id, name, price, rate_limit FROM profiles WHERE type=%s ORDER BY price ASC",
            (p_type,), fetch=True
        ) or []

    from flask import jsonify
    return jsonify([dict(p) for p in profiles])

@profiles_bp.route('/import', methods=['GET', 'POST'])
def import_excel():
    if not session.get('logged_in'):
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file', 'error')
            return redirect(request.url)
            
        if file and file.filename.endswith('.xlsx'):
            import os
            from openpyxl import load_workbook
            from web.app import UPLOAD_FOLDER
            
            filepath = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filepath)
            
            try:
                wb = load_workbook(filepath)
                sheet = wb.active
                
                count = 0
                # Header: Name, Rate Limit, Price, Validity, Unit, Type, Router
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    name, rate, price, validity, unit, p_type, router_name = row
                    if not name: continue
                    
                    router_id = None
                    if router_name:
                        r = execute_query("SELECT id FROM routers WHERE name=%s", (router_name,), fetch_one=True)
                        if r: router_id = r['id']

                    execute_query(
                        "INSERT INTO profiles (name, rate_limit, price, validity, validity_unit, type, router_id) "
                        "VALUES (%s, %s, %s, %s, %s, %s, %s) "
                        "ON DUPLICATE KEY UPDATE rate_limit=%s, price=%s, validity=%s, validity_unit=%s, type=%s, router_id=%s",
                        (name, rate or '1M/1M', price or 0, validity or 0, unit or 'hours', p_type or 'pppoe', router_id,
                         rate or '1M/1M', price or 0, validity or 0, unit or 'hours', p_type or 'pppoe', router_id)
                    )
                    count += 1
                
                flash(f'Berhasil mengimport {count} Paket Internet.', 'success')
                return redirect(url_for('profiles.index'))
            except Exception as e:
                flash(f'Error reading excel: {e}', 'error')
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    
    return render_template('profiles/import.html')
